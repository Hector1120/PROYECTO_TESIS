import random
import re
import string
import json
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import Usuario  
from django.utils import timezone
from datetime import datetime, timedelta, time
from .models import Horario, BloqueHorario, Area, Asignatura 
from .models import PermisoHorario, Periodo, PermisoDocente, Asesoria, AsesoriaBloque
from django.core.exceptions import ObjectDoesNotExist
from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
import csv
from django.db.models import Q
from django.contrib.auth.decorators import login_required


@csrf_exempt
def registrar_horario(request):
    if request.method == "POST":
        try:
            if not request.user.is_authenticated:
                return JsonResponse({"mensaje": "No autorizado"}, status=401)
            
            # Verificar que el usuario sea docente
            if request.user.rol != "Docente":
                return JsonResponse({"mensaje": "Solo los docentes pueden registrar horarios"}, status=403)
            
            try:
                permiso = PermisoDocente.objects.get(docente=request.user)
                if not permiso.puede_crear_horarios:
                    return JsonResponse({"mensaje": "No tiene permiso para crear horarios"}, status=403)
            except PermisoDocente.DoesNotExist:
                return JsonResponse({"mensaje": "No tiene permiso para crear horarios"}, status=403)
            
            data = json.loads(request.body)
            periodo = data.get("periodo", "2025-1")
            dia = data.get("dia")
            hora_inicio_str = data.get("hora_inicio")
            hora_fin_str = data.get("hora_fin")
            lugar = data.get("lugar")  # Nuevo campo para lugar
            docente_id = request.user.id
            
            # Validar que el lugar no esté vacío
            if not lugar:
                return JsonResponse({"mensaje": "El lugar de encuentro es obligatorio"}, status=400)
                
            # Validar día de la semana
            dias_validos = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
            if dia not in dias_validos:
                return JsonResponse({"mensaje": "Día no válido. Debe ser de lunes a viernes."}, status=400)
            # Convertir horas de string a objeto time
            hora_inicio = datetime.strptime(hora_inicio_str, "%H:%M").time()
            hora_fin = datetime.strptime(hora_fin_str, "%H:%M").time()
            # Definir el rango de horas permitido
            hora_min = time(7, 0)  # 7:00 AM
            hora_max = time(20, 0)  # 8:00 PM
            # Validar el rango de horas
            if hora_inicio < hora_min or hora_inicio > hora_max:
                return JsonResponse({"mensaje": "La hora de inicio debe estar entre 7:00 AM y 8:00 PM"}, status=400)
            if hora_fin < hora_min or hora_fin > hora_max:
                return JsonResponse({"mensaje": "La hora de fin debe estar entre 7:00 AM y 8:00 PM"}, status=400)
            if hora_inicio >= hora_fin:
                return JsonResponse({"mensaje": "La hora de inicio debe ser anterior a la hora de fin"}, status=400)

            # Validar que la diferencia sea múltiplo de 15 minutos
            inicio_minutos = hora_inicio.hour * 60 + hora_inicio.minute
            fin_minutos = hora_fin.hour * 60 + hora_fin.minute
            diferencia_minutos = fin_minutos - inicio_minutos
            
            if diferencia_minutos < 15:
                return JsonResponse({"mensaje": "La duración mínima debe ser de 15 minutos"}, status=400)
            
            if diferencia_minutos % 15 != 0:
                return JsonResponse({"mensaje": "La duración debe ser múltiplo de 15 minutos (15, 30, 45, 60 minutos, etc.)"}, status=400)
            
            # Obtener el periodo vigente desde el modelo
            periodo_obj = Periodo.objects.filter(fecha_fin__gte=timezone.now().date(), activo=True).order_by('fecha_inicio').first()
            if not periodo_obj:
                return JsonResponse({"mensaje": "No hay periodo vigente"}, status=400)

            # Extraer fecha de inicio y fin desde el periodo vigente
            fecha_inicio = periodo_obj.fecha_inicio
            fecha_fin = periodo_obj.fecha_fin

            # Calcular todos los días que corresponden al día seleccionado dentro del rango de fechas
            dias_semana = {
                "Lunes": 0,
                "Martes": 1,
                "Miércoles": 2,
                "Jueves": 3,
                "Viernes": 4,
            }
            dia_seleccionado = dias_semana[dia]
            fechas_a_registrar = []
            current_date = fecha_inicio
            while current_date <= fecha_fin:
                if current_date.weekday() == dia_seleccionado:
                    fechas_a_registrar.append(current_date)
                current_date += timedelta(days=1)
            # Registrar horarios para cada fecha calculada
            for fecha in fechas_a_registrar:
                # Crear nuevo horario sin verificar si ya existe
                horario = Horario.objects.create(
                    docente_id=docente_id,
                    periodo=periodo,
                    dia=dia,
                    fecha=fecha,  # Registrar la fecha
                    hora_inicio=hora_inicio,
                    hora_fin=hora_fin
                )
                # Generar bloques de 15 minutos
                current_time = datetime.combine(fecha, hora_inicio)
                end_time = datetime.combine(fecha, hora_fin)

                while current_time < end_time:
                    BloqueHorario.objects.create(
                        horario=horario,
                        fecha=fecha,
                        hora_inicio=current_time.time(),
                        hora_fin=(current_time + timedelta(minutes=15)).time(),
                        lugar=lugar  # Añadir el lugar a cada bloque
                    )
                    current_time += timedelta(minutes=15)  # Incrementar 15 minutos
                
            return JsonResponse({"mensaje": "Horarios registrados con éxito"}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Datos inválidos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)


def eliminar_horarios_fuera_periodo():
    """Elimina los horarios que estén fuera del periodo vigente"""
    try:
        # Obtener el periodo vigente
        periodo_vigente = Periodo.objects.filter(activo=True).first()
        
        if not periodo_vigente:
            return False
            
        # Eliminar horarios fuera del rango de fechas del periodo
        Horario.objects.filter(
            fecha__lt=periodo_vigente.fecha_inicio
        ).delete()
        
        Horario.objects.filter(
            fecha__gt=periodo_vigente.fecha_fin
        ).delete()
        
        return True
    except Exception as e:
        print(f"Error al eliminar horarios fuera del periodo: {str(e)}")
        return False



@csrf_exempt
def limpiar_horarios_endpoint(request):
    """Endpoint para eliminar horarios pasados y fuera del periodo vigente"""
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
        
    if request.user.rol not in ["Administrador", "Director"]:
        return JsonResponse({"mensaje": "No tiene permisos suficientes"}, status=403)
        
    # Eliminar horarios fuera del periodo
    resultado = eliminar_horarios_fuera_periodo()
    
    if resultado:
        return JsonResponse({
            "mensaje": "Horarios fuera de periodo eliminados correctamente"
        }, status=200)
    else:
        return JsonResponse({
            "mensaje": "No se encontró un periodo vigente activo"
        }, status=400)
def generar_contraseña(longitud=8):
    """Genera una contraseña aleatoria con letras y números."""
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(longitud))

@csrf_exempt
def recuperar_contraseña(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            correo = data.get("correo")
            
            usuario = Usuario.objects.filter(correo=correo).first()
            if usuario:
                nueva_contraseña = generar_contraseña()
                usuario.set_password(nueva_contraseña)
                usuario.save()
                
                send_mail(
                    "Recuperación de Contraseña - Sistema de Asesorías CESMAG",
                    f"Tu nueva contraseña es: {nueva_contraseña}\nPor favor, cambia tu contraseña después de iniciar sesión.",
                    "lizcanoagudelo2@gmail.com",
                    [correo],
                    fail_silently=False,
                )
                return JsonResponse({"mensaje": "Se ha enviado una nueva contraseña a tu correo."}, status=200)
            else:
                return JsonResponse({"mensaje": "El correo ingresado no está registrado."}, status=404)
        
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos."}, status=400)
    
    return JsonResponse({"mensaje": "Usa el método POST para recuperar la contraseña."}, status=405)

def extraer_rol(nombre):
    """Extrae el rol del nombre, asumiendo que viene entre corchetes al final."""
    if "[Estudiante]" in nombre:
        return "Estudiante"
    elif "[Docente]" in nombre:
        return "Docente"
    return "Estudiante"  # Valor por defecto

def enviar_correo(request):
    usuarios = {
        "hylizcano.1799@unicesmag.edu.co": "Hector Yurbrainer Lizcano Agudelo [Estudiante]",
        "hmsolarte.5416@unicesmag.edu.co": "Hector Mauricio Solarte Muñoz [Docente]",
        # Agrega más usuarios aquí
    }

    contraseñas = {}

    for correo, nombre in usuarios.items():
        contraseña = generar_contraseña()
        rol = extraer_rol(nombre)  # Extraer rol desde el nombre

        # Buscar si el usuario ya existe
        usuario, creado = Usuario.objects.get_or_create(correo=correo)
        
        # Si el usuario ya existía, actualizar el rol y la contraseña
        usuario.rol = rol
        usuario.set_password(contraseña)  # Guarda la contraseña encriptada
        usuario.save()

        contraseñas[correo] = contraseña  # Guardamos en el diccionario
        
        send_mail(
            "Acceso al Sistema de Asesorías CESMAG",
            f"Tu contraseña por defecto es: {contraseña}\nTu rol es: {rol}",
            "lizcanoagudelo2@gmail.com",
            [correo],
            fail_silently=False,
        )

    return JsonResponse({"mensaje": "Correos enviados y usuarios actualizados", "contraseñas": contraseñas})

# Modificar la función login_usuario en views.py
@csrf_exempt
def login_usuario(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            correo = data.get("correo")
            contraseña = data.get("contraseña")
            usuario = Usuario.objects.filter(correo=correo).first()
            
            # Verificar si el usuario existe y la contraseña es correcta
            if usuario and usuario.check_password(contraseña):
                # Verificar si la cuenta está activa
                if not usuario.is_active:
                    return JsonResponse({"mensaje": "La cuenta está inactiva. Contacte al administrador."}, status=403)
                
                # Si la cuenta está activa, proceder con el inicio de sesión
                from django.contrib.auth import login
                login(request, usuario)
                
                # Incluir subtipo_director en la respuesta
                return JsonResponse({
                    "mensaje": "Inicio de sesión exitoso", 
                    "rol": usuario.rol,
                    "nombre_usuario": usuario.nombre_usuario,
                    "correo": usuario.correo,
                    "subtipo_director": usuario.subtipo_director if hasattr(usuario, 'subtipo_director') else ""
                }, status=200)
            else:
                return JsonResponse({"mensaje": "El correo o la contraseña no son correctos"}, status=401)
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
    return JsonResponse({"mensaje": "Usa el método POST para iniciar sesión"}, status=405)

# También asegúrate de incluir el subtipo_director en verificar_sesion
@csrf_exempt
def verificar_sesion(request):
    # Verifica si hay una sesión activa
    if request.user.is_authenticated:
        return JsonResponse({
            "autenticado": True,
            "usuario": {
                "correo": request.user.correo,
                "rol": request.user.rol,
                "nombre_usuario": request.user.nombre_usuario,
                "subtipo_director": request.user.subtipo_director if hasattr(request.user, 'subtipo_director') else ""
            }
        })
    else:
        return JsonResponse({"autenticado": False}, status=401)
    
# Añadir a views.py
@csrf_exempt
def logout_usuario(request):
    from django.contrib.auth import logout
    logout(request)
    return JsonResponse({"mensaje": "Sesión cerrada correctamente"})


@csrf_exempt
def obtener_horarios(request):
    if request.method == "GET":
        try:
            if not request.user.is_authenticated:
                return JsonResponse({"mensaje": "No autorizado"}, status=401)
            
            # Obtener el docente actual
            docente_id = request.user.id
            
            # Obtener el periodo vigente
            periodo_vigente = Periodo.objects.filter(activo=True).first()
            
            if not periodo_vigente:
                return JsonResponse({"mensaje": "No hay un periodo vigente configurado"}, status=400)
            
            # Obtener la fecha y hora actual
            ahora = timezone.localtime(timezone.now())
            fecha_actual = ahora.date()
            hora_actual = ahora.time()
            
            # Buscar todos los bloques de horario del docente dentro del periodo vigente
            horarios = Horario.objects.filter(
                docente_id=docente_id,
                fecha__gte=periodo_vigente.fecha_inicio,
                fecha__lte=periodo_vigente.fecha_fin
            )
            
            bloques_horario = []
            
            for horario in horarios:
                # Obtener bloques asociados a este horario
                bloques = BloqueHorario.objects.filter(horario=horario)
                
                for bloque in bloques:
                    # Filtrar para mostrar solo horarios futuros
                    bloque_fecha = bloque.fecha
                    bloque_hora_fin = bloque.hora_fin
                    
                    # Incluir solo horarios futuros (fecha posterior a hoy o del mismo día con hora posterior a la actual)
                    if bloque_fecha > fecha_actual or (bloque_fecha == fecha_actual and bloque_hora_fin > hora_actual):
                        bloques_horario.append({
                            "horario_id": horario.id,
                            "bloque_id": bloque.id,
                            "dia": horario.dia,
                            "fecha": bloque.fecha.strftime("%Y-%m-%d"),
                            "hora_inicio": bloque.hora_inicio.strftime("%H:%M"),
                            "hora_fin": bloque.hora_fin.strftime("%H:%M"),
                            "lugar": bloque.lugar,  # Incluir el lugar en la respuesta
                            "periodo": horario.periodo
                        })
            
            return JsonResponse({
                "bloques": bloques_horario,
                "periodo": periodo_vigente.codigo,
                "fecha_inicio": periodo_vigente.fecha_inicio.strftime("%Y-%m-%d"),
                "fecha_fin": periodo_vigente.fecha_fin.strftime("%Y-%m-%d")
            }, status=200)
        
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def obtener_horarios_docentes(request):
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol != "Director":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            # Obtener el subtipo del director
            subtipo_director = request.user.subtipo_director
            subtipo_docente_permitido = None
            
            # Mapear subtipo de director a subtipo de docente
            if subtipo_director == "Director de Ing.Sistemas":
                subtipo_docente_permitido = "Docente de Ing.Sistemas"
            elif subtipo_director == "Director de Ing.Electrónica":
                subtipo_docente_permitido = "Docente de Ing.Electrónica"
            elif subtipo_director == "Director de Humanidades":
                subtipo_docente_permitido = "Docente de Humanidades"
            elif subtipo_director == "Director de Idiomas":
                subtipo_docente_permitido = "Docente de Idiomas"
            elif subtipo_director == "Director de Ciencias Básicas":
                subtipo_docente_permitido = "Docente de Ciencias Básicas"
            
            # Obtener todos los docentes del subtipo correspondiente
            docentes = Usuario.objects.filter(
                rol="Docente", 
                subtipo_docente=subtipo_docente_permitido
            )
            
            # Obtener la fecha y hora actual para filtrar horarios pasados
            ahora = timezone.localtime(timezone.now())
            fecha_actual = ahora.date()
            hora_actual = ahora.time()
            
            # Obtener los horarios de cada docente
            resultados = []
            
            for docente in docentes:
                # Verificar permisos para crear horarios
                try:
                    permiso_docente = PermisoDocente.objects.get(docente=docente)
                    puede_crear_horarios = permiso_docente.puede_crear_horarios
                except PermisoDocente.DoesNotExist:
                    puede_crear_horarios = False
                
                # Obtener todos los horarios del docente
                horarios = Horario.objects.filter(docente_id=docente.id)
                horarios_docente = []
                
                for horario in horarios:
                    # Solo incluir horarios futuros (fecha posterior a hoy o del mismo día con hora posterior a la actual)
                    if horario.fecha > fecha_actual or (horario.fecha == fecha_actual and horario.hora_fin > hora_actual):
                        # Verificar si este horario tiene permisos especiales
                        try:
                            permiso = PermisoHorario.objects.get(horario=horario)
                            puede_editar = permiso.puede_editar
                            puede_eliminar = permiso.puede_eliminar
                        except ObjectDoesNotExist:
                            puede_editar = False
                            puede_eliminar = False
                        
                        horarios_docente.append({
                            "id": horario.id,
                            "dia": horario.dia,
                            "fecha": horario.fecha.strftime("%Y-%m-%d"),
                            "hora_inicio": horario.hora_inicio.strftime("%H:%M"),
                            "hora_fin": horario.hora_fin.strftime("%H:%M"),
                            "periodo": horario.periodo,
                            "puede_editar": puede_editar,
                            "puede_eliminar": puede_eliminar
                        })
                
                # Agregamos a todos los docentes, tengan o no horarios
                resultados.append({
                    "docente_id": docente.id,
                    "correo": docente.correo,
                    "subtipo_docente": docente.subtipo_docente,
                    "puede_crear_horarios": puede_crear_horarios,
                    "horarios": horarios_docente
                })
            
            return JsonResponse({"docentes_horarios": resultados}, status=200)
            
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)
@csrf_exempt
def actualizar_permisos_docente(request):
    if request.method == "POST":
        if not request.user.is_authenticated or request.user.rol != "Director":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            data = json.loads(request.body)
            docente_id = data.get("docente_id")
            puede_crear_horarios = data.get("puede_crear_horarios", False)
            
            # Verificar que el docente pertenece al área del director
            docente = Usuario.objects.get(id=docente_id, rol="Docente")
            
            # Mapear subtipo de director a subtipo de docente
            subtipo_director = request.user.subtipo_director
            subtipo_docente_permitido = None
            
            if subtipo_director == "Director de Ing.Sistemas":
                subtipo_docente_permitido = "Docente de Ing.Sistemas"
            elif subtipo_director == "Director de Ing.Electrónica":
                subtipo_docente_permitido = "Docente de Ing.Electrónica"
            elif subtipo_director == "Director de Humanidades":
                subtipo_docente_permitido = "Docente de Humanidades"
            elif subtipo_director == "Director de Idiomas":
                subtipo_docente_permitido = "Docente de Idiomas"
            elif subtipo_director == "Director de Ciencias Básicas":
                subtipo_docente_permitido = "Docente de Ciencias Básicas"
            
            # Verificar que el docente pertenece al área del director
            if docente.subtipo_docente != subtipo_docente_permitido:
                return JsonResponse({"mensaje": "No tiene permisos sobre este docente"}, status=403)
            
            # Crear o actualizar los permisos
            permiso, created = PermisoDocente.objects.get_or_create(
                docente=docente,
                defaults={"puede_crear_horarios": puede_crear_horarios}
            )
            
            if not created:
                permiso.puede_crear_horarios = puede_crear_horarios
                permiso.save()
            
            return JsonResponse({
                "mensaje": "Permisos actualizados correctamente",
                "docente_id": docente_id,
                "puede_crear_horarios": puede_crear_horarios
            }, status=200)
            
        except Usuario.DoesNotExist:
            return JsonResponse({"mensaje": "Docente no encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def actualizar_permisos_horario(request):
    if request.method == "POST":
        if not request.user.is_authenticated or request.user.rol != "Director":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            data = json.loads(request.body)
            horario_id = data.get("horario_id")
            puede_editar = data.get("puede_editar", False)
            puede_eliminar = data.get("puede_eliminar", False)
            
            horario = Horario.objects.get(id=horario_id)
            
            # Crear o actualizar los permisos
            permiso, created = PermisoHorario.objects.get_or_create(
                horario=horario,
                defaults={
                    "puede_editar": puede_editar,
                    "puede_eliminar": puede_eliminar
                }
            )
            
            if not created:
                permiso.puede_editar = puede_editar
                permiso.puede_eliminar = puede_eliminar
                permiso.save()
            
            return JsonResponse({
                "mensaje": "Permisos actualizados correctamente",
                "horario_id": horario_id,
                "puede_editar": puede_editar,
                "puede_eliminar": puede_eliminar
            }, status=200)
            
        except ObjectDoesNotExist:
            return JsonResponse({"mensaje": "Horario no encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def verificar_permisos_horario(request, horario_id):
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol != "Docente":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            # Verificar que el horario pertenezca al docente
            horario = Horario.objects.get(id=horario_id, docente_id=request.user.id)
            
            # Buscar los permisos asociados
            try:
                permiso = PermisoHorario.objects.get(horario=horario)
                return JsonResponse({
                    "puede_editar": permiso.puede_editar,
                    "puede_eliminar": permiso.puede_eliminar
                }, status=200)
            except ObjectDoesNotExist:
                return JsonResponse({
                    "puede_editar": False,
                    "puede_eliminar": False
                }, status=200)
                
        except ObjectDoesNotExist:
            return JsonResponse({"mensaje": "Horario no encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def verificar_permiso_crear_horarios(request):
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol != "Docente":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            # Buscar los permisos asociados
            try:
                permiso = PermisoDocente.objects.get(docente=request.user)
                return JsonResponse({
                    "puede_crear_horarios": permiso.puede_crear_horarios
                }, status=200)
            except PermisoDocente.DoesNotExist:
                return JsonResponse({
                    "puede_crear_horarios": False
                }, status=200)
                
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def editar_horario(request):
    if request.method == "PUT":
        try:
            if not request.user.is_authenticated or request.user.rol != "Docente":
                return JsonResponse({"mensaje": "No autorizado"}, status=401)
            
            data = json.loads(request.body)
            horario_id = data.get("horario_id")
            dia = data.get("dia")
            fecha_str = data.get("fecha")
            hora_inicio_str = data.get("hora_inicio")
            hora_fin_str = data.get("hora_fin")
            lugar = data.get("lugar")
            
            # Validar que el lugar no esté vacío
            if not lugar:
                return JsonResponse({"mensaje": "El lugar de encuentro es obligatorio"}, status=400)
            
            # Verificar que el horario pertenezca al docente
            try:
                horario = Horario.objects.get(id=horario_id, docente_id=request.user.id)
            except ObjectDoesNotExist:
                return JsonResponse({"mensaje": "Horario no encontrado"}, status=404)
            
            # Verificar permisos de edición
            try:
                permiso = PermisoHorario.objects.get(horario=horario)
                puede_editar = permiso.puede_editar
            except ObjectDoesNotExist:
                puede_editar = False
            
            # Si no tiene permisos de edición, solo actualizar el lugar
            if not puede_editar:
                # Actualizar solo el lugar en los bloques de horario
                BloqueHorario.objects.filter(horario=horario).update(lugar=lugar)
                return JsonResponse({"mensaje": "Lugar actualizado con éxito"}, status=200)
            
            # Si tiene permisos de edición, proceder con la validación completa
            dias_validos = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
            if dia not in dias_validos:
                return JsonResponse({"mensaje": "Día no válido. Debe ser de lunes a viernes."}, status=400)
            
            # Validar y convertir la fecha
            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({"mensaje": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)
            
            # Validar que el día de la semana coincida con la fecha
            weekday_map = {
                0: 'Lunes',
                1: 'Martes',
                2: 'Miércoles',
                3: 'Jueves',
                4: 'Viernes',
                5: 'Sábado',
                6: 'Domingo'
            }
            dia_semana_python = fecha.weekday()  # Python's weekday() returns 0 for Monday
            dia_esperado = weekday_map[dia_semana_python]
            if dia != dia_esperado:
                return JsonResponse({
                    "mensaje": f"El día seleccionado ({dia}) no coincide con el día de la fecha seleccionada ({dia_esperado})"
                }, status=400)
            
            hora_inicio = datetime.strptime(hora_inicio_str, "%H:%M").time()
            hora_fin = datetime.strptime(hora_fin_str, "%H:%M").time()
            
            # Definir el rango de horas permitido
            hora_min = time(7, 0)  # 7:00 AM
            hora_max = time(20, 0)  # 8:00 PM
            
            # Validar el rango de horas
            if hora_inicio < hora_min or hora_inicio > hora_max:
                return JsonResponse({"mensaje": "La hora de inicio debe estar entre 7:00 AM y 8:00 PM"}, status=400)
            if hora_fin < hora_min or hora_fin > hora_max:
                return JsonResponse({"mensaje": "La hora de fin debe estar entre 7:00 AM y 8:00 PM"}, status=400)
            if hora_inicio >= hora_fin:
                return JsonResponse({"mensaje": "La hora de inicio debe ser anterior a la hora de fin"}, status=400)

            # Actualizar horario
            horario.dia = dia
            horario.fecha = fecha
            horario.hora_inicio = hora_inicio
            horario.hora_fin = hora_fin
            horario.save()
            
            # Actualizar bloques de horario
            # Primero eliminar los bloques existentes
            BloqueHorario.objects.filter(horario=horario).delete()
            
            # Crear nuevos bloques
            current_time = datetime.combine(fecha, hora_inicio)
            end_time = datetime.combine(fecha, hora_fin)
            
            while current_time < end_time:
                BloqueHorario.objects.create(
                    horario=horario,
                    fecha=fecha,
                    hora_inicio=current_time.time(),
                    hora_fin=(current_time + timedelta(minutes=15)).time(),
                    lugar=lugar  # Añadir el lugar a cada bloque
                )
                current_time += timedelta(minutes=15)
            
            return JsonResponse({"mensaje": "Horario actualizado con éxito"}, status=200)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Datos inválidos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def eliminar_horario(request):
    if request.method == "DELETE":
        try:
            if not request.user.is_authenticated or request.user.rol != "Docente":
                return JsonResponse({"mensaje": "No autorizado"}, status=401)
            
            data = json.loads(request.body)
            horario_id = data.get("horario_id")
            
            # Verificar que el horario pertenezca al docente
            try:
                horario = Horario.objects.get(id=horario_id, docente_id=request.user.id)
            except ObjectDoesNotExist:
                return JsonResponse({"mensaje": "Horario no encontrado"}, status=404)
            
            # Verificar permisos de eliminación
            try:
                permiso = PermisoHorario.objects.get(horario=horario)
                if not permiso.puede_eliminar:
                    return JsonResponse({"mensaje": "No tiene permiso para eliminar este horario"}, status=403)
            except ObjectDoesNotExist:
                return JsonResponse({"mensaje": "No tiene permiso para eliminar este horario"}, status=403)
            
            # Eliminar horario y sus bloques (la eliminación en cascada se encargará de los bloques)
            horario.delete()
            
            return JsonResponse({"mensaje": "Horario eliminado con éxito"}, status=200)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Datos inválidos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def listar_usuarios(request):
    """Obtiene todos los usuarios del sistema"""
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            usuarios = Usuario.objects.all()
            data = []
            
            for usuario in usuarios:
                usuario_data = {
                    "id": usuario.id,
                    "nombre_usuario": usuario.nombre_usuario,  # Incluimos el nombre del usuario
                    "correo": usuario.correo,
                    "rol": usuario.rol,
                    "is_active": usuario.is_active,
                    "is_staff": usuario.is_staff,
                    "subtipo_director": usuario.subtipo_director if hasattr(usuario, 'subtipo_director') else "",
                    "subtipo_docente": usuario.subtipo_docente if hasattr(usuario, 'subtipo_docente') else ""
                }
                
                # Añadir campos de estudiante si el usuario es un estudiante
                if usuario.rol == "Estudiante":
                    usuario_data["semestre"] = usuario.semestre
                    usuario_data["grupo"] = usuario.grupo
                
                data.append(usuario_data)
            
            return JsonResponse(data, safe=False)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def crear_usuario(request):
    """Crea un nuevo usuario y envía la contraseña por correo"""
    if request.method == "POST":
        if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
       
        try:
            data = json.loads(request.body)
           
            # Validar que se envíen los campos requeridos
            if "correo" not in data:
                return JsonResponse({"mensaje": "El correo es requerido"}, status=400)
            
            # Validar que se envíe el nombre (campo obligatorio)
            if "nombre_usuario" not in data:
                return JsonResponse({"mensaje": "El nombre es requerido"}, status=400)
           
            # Verificar si ya existe un usuario con el mismo correo
            if Usuario.objects.filter(correo=data["correo"]).exists():
                return JsonResponse({"mensaje": "Ya existe un usuario con este correo"}, status=400)
            
            # Si es un director, validar que se incluya el subtipo y que no esté duplicado
            if data.get("rol") == "Director":
                if not data.get("subtipo_director"):
                    return JsonResponse({"mensaje": "El tipo de director es requerido"}, status=400)
                
                # Verificar si ya existe un usuario con ese subtipo de director
                if Usuario.objects.filter(rol="Director", subtipo_director=data.get("subtipo_director")).exists():
                    return JsonResponse({"mensaje": f"Ya existe un '{data.get('subtipo_director')}' registrado"}, status=400)
            
            # Si es un docente, validar que se incluya el subtipo
            if data.get("rol") == "Docente" and not data.get("subtipo_docente"):
                return JsonResponse({"mensaje": "El tipo de docente es requerido"}, status=400)
           
            # Generar contraseña aleatoria
            contraseña = generar_contraseña()
           
            # Determinar si el usuario es superusuario basado en su rol
            es_superusuario = data.get("rol") in ["Administrador", "Director"]
            
            # Crear el nuevo usuario
            nuevo_usuario = Usuario(
                nombre_usuario=data["nombre_usuario"],  # Incluimos el nombre del usuario
                correo=data["correo"],
                rol=data.get("rol", "Estudiante"),
                subtipo_director=data.get("subtipo_director", ""),
                subtipo_docente=data.get("subtipo_docente", ""),
                is_active=data.get("is_active", True),
                is_staff=es_superusuario,
                is_superuser=es_superusuario
            )
            
            # Asignar semestre y grupo si es estudiante
            if data.get("rol") == "Estudiante" or data.get("rol") is None:
                nuevo_usuario.semestre = data.get("semestre", "")
                nuevo_usuario.grupo = data.get("grupo", "")
                
            nuevo_usuario.set_password(contraseña)
            nuevo_usuario.save()
            
            # Preparar el mensaje de correo
            mensaje_correo = f"Hola {nuevo_usuario.nombre_usuario},\n\nTú contraseña de acceso es: {contraseña}\n\nY tu rol es: {nuevo_usuario.rol}"
            
            # Añadir información específica según el rol
            if nuevo_usuario.rol == "Director" and nuevo_usuario.subtipo_director:
                mensaje_correo += f" - {nuevo_usuario.subtipo_director}"
            elif nuevo_usuario.rol == "Docente" and nuevo_usuario.subtipo_docente:
                mensaje_correo += f" - {nuevo_usuario.subtipo_docente}"
            elif nuevo_usuario.rol == "Estudiante":
                mensaje_correo += f"\nSemestre: {nuevo_usuario.semestre or 'No asignado'}"
                mensaje_correo += f"\nGrupo: {nuevo_usuario.grupo or 'No asignado'}"
                
            mensaje_correo += "."
            
            # Enviar correo con la contraseña
            send_mail(
                "Acceso al Sistema de Asesorías CESMAG",
                mensaje_correo,
                "lizcanoagudelo2@gmail.com",
                [data["correo"]],
                fail_silently=False,
            )
           
            return JsonResponse({
                "mensaje": "Usuario creado exitosamente. Contraseña enviada por correo.",
                "id": nuevo_usuario.id
            }, status=201)
           
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
   
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def detalle_usuario(request, usuario_id):
    """Obtiene, actualiza o elimina un usuario específico"""
    if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
        return JsonResponse({"mensaje": "No autorizado"}, status=403)
    
    try:
        usuario = Usuario.objects.get(id=usuario_id)
    except ObjectDoesNotExist:
        return JsonResponse({"mensaje": "Usuario no encontrado"}, status=404)
    
    # GET: Obtener detalle de un usuario
    if request.method == "GET":
        data = {
            "id": usuario.id,
            "nombre_usuario": usuario.nombre_usuario,  # Incluimos el nombre del usuario
            "correo": usuario.correo,
            "password": usuario.password,
            "rol": usuario.rol,
            "subtipo_director": usuario.subtipo_director if hasattr(usuario, 'subtipo_director') else "",
            "subtipo_docente": usuario.subtipo_docente if hasattr(usuario, 'subtipo_docente') else "",
            "is_active": usuario.is_active,
            "is_staff": usuario.is_staff
        }
        
        # Añadir campos de estudiante si el usuario es un estudiante
        if usuario.rol == "Estudiante":
            data["semestre"] = usuario.semestre
            data["grupo"] = usuario.grupo
            
        return JsonResponse(data)
    
    # PUT: Actualizar un usuario
    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            
            # Guardar valores antiguos para comparación
            old_rol = usuario.rol
            old_semestre = usuario.semestre if usuario.rol == "Estudiante" else None
            
            # Actualizar datos del usuario
            if "nombre_usuario" in data:
                usuario.nombre_usuario = data["nombre_usuario"]
                
            if "correo" in data:
                # Verificar si el nuevo correo ya existe y no es del usuario actual
                if Usuario.objects.filter(correo=data["correo"]).exclude(id=usuario_id).exists():
                    return JsonResponse({"mensaje": "El correo ya está en uso por otro usuario"}, status=400)
                usuario.correo = data["correo"]
            
            if "rol" in data:
                new_rol = data["rol"]
                
                # Validar que el rol sea válido
                if new_rol not in dict(Usuario.ROLES):
                    return JsonResponse({"mensaje": f"Rol no válido. Opciones: {', '.join(dict(Usuario.ROLES).keys())}"}, status=400)
                
                # Si cambia a Director, verificar que tenga subtipo_director
                if new_rol == "Director" and "subtipo_director" in data:
                    # Verificar que el subtipo no esté ya asignado a otro director
                    if Usuario.objects.filter(
                        rol="Director", 
                        subtipo_director=data["subtipo_director"]
                    ).exclude(id=usuario_id).exists():
                        return JsonResponse({"mensaje": f"Ya existe un '{data['subtipo_director']}' registrado"}, status=400)
                    
                    # Asignar el subtipo de director
                    usuario.subtipo_director = data["subtipo_director"]
                    usuario.subtipo_docente = ""  # Limpiar el subtipo de docente
                    usuario.semestre = ""  # Limpiar el semestre
                    usuario.grupo = ""     # Limpiar el grupo
                
                # Si cambia a Docente, validar el subtipo_docente
                elif new_rol == "Docente" and "subtipo_docente" in data:
                    usuario.subtipo_docente = data["subtipo_docente"]
                    usuario.subtipo_director = ""  # Limpiar el subtipo de director
                    usuario.semestre = ""  # Limpiar el semestre
                    usuario.grupo = ""     # Limpiar el grupo
                
                # Si cambia a Estudiante, permitir actualizar semestre y grupo
                elif new_rol == "Estudiante":
                    usuario.subtipo_director = ""  # Limpiar el subtipo de director
                    usuario.subtipo_docente = ""   # Limpiar el subtipo de docente
                    
                    # Actualizar semestre y grupo si se proporcionan en los datos
                    if "semestre" in data:
                        usuario.semestre = data["semestre"]
                    if "grupo" in data:
                        usuario.grupo = data["grupo"]
                        
                # Si cambia de Director/Docente a otro rol, limpiar los subtipos
                else:
                    if old_rol == "Director":
                        usuario.subtipo_director = ""
                    elif old_rol == "Docente":
                        usuario.subtipo_docente = ""
                    elif old_rol == "Estudiante":
                        usuario.semestre = ""
                        usuario.grupo = ""
                
                usuario.rol = new_rol
                # Actualizar is_staff si el rol es Administrador o Director
                usuario.is_staff = True if new_rol in ["Administrador", "Director"] else False
                usuario.is_superuser = True if new_rol in ["Administrador", "Director"] else False
            
            # Si no cambia el rol pero sí el subtipo (siendo Director)
            elif usuario.rol == "Director" and "subtipo_director" in data:
                # Verificar que el subtipo no esté ya asignado a otro director
                if Usuario.objects.filter(
                    rol="Director", 
                    subtipo_director=data["subtipo_director"]
                ).exclude(id=usuario_id).exists():
                    return JsonResponse({"mensaje": f"Ya existe un '{data['subtipo_director']}' registrado"}, status=400)
                
                usuario.subtipo_director = data["subtipo_director"]
            
            # Si no cambia el rol pero sí el subtipo (siendo Docente)
            elif usuario.rol == "Docente" and "subtipo_docente" in data:
                usuario.subtipo_docente = data["subtipo_docente"]
            
            # Si es estudiante y se actualizan semestre o grupo
            elif usuario.rol == "Estudiante":
                new_semestre = data.get("semestre", usuario.semestre)
                if "semestre" in data:
                    usuario.semestre = data["semestre"]
                if "grupo" in data:
                    usuario.grupo = data["grupo"]
            
            if "is_active" in data:
                usuario.is_active = data["is_active"]
            
            if "password" in data and data["password"]:
                usuario.set_password(data["password"])
            
            # Guardar el usuario primero para tener los cambios actualizados
            usuario.save()
            
            # Eliminar asesorías si:
            # 1. Cambió de rol y ya no es estudiante
            # 2. Sigue siendo estudiante pero cambió de semestre
            if (old_rol == "Estudiante" and usuario.rol != "Estudiante") or \
               (usuario.rol == "Estudiante" and "semestre" in data and old_semestre != data["semestre"]):
                
                # Obtener todas las asesorías del estudiante
                asesorias = Asesoria.objects.filter(estudiante=usuario)
                
                # Notificar a los docentes antes de eliminar
                for asesoria in asesorias:
                    try:
                        send_mail(
                            'Asesoría cancelada por cambio de estado del estudiante',
                            f'La asesoría programada con el estudiante {usuario.nombre_usuario} ({usuario.correo}) ha sido cancelada debido a un cambio en su rol/semestre.',
                            'sistema@example.com',
                            [asesoria.docente.correo],
                            fail_silently=True,
                        )
                    except Exception as e:
                        print(f"Error al enviar correo de notificación: {e}")
                
                # Eliminar todas las asesorías
                asesorias.delete()
            
            return JsonResponse({"mensaje": "Usuario actualizado exitosamente"})
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # DELETE: Eliminar un usuario
    elif request.method == "DELETE":
        # Notificar y eliminar asesorías si el usuario es estudiante
        if usuario.rol == "Estudiante":
            asesorias = Asesoria.objects.filter(estudiante=usuario)
            for asesoria in asesorias:
                try:
                    send_mail(
                        'Asesoría cancelada por eliminación de estudiante',
                        f'La asesoría programada con el estudiante {usuario.nombre_usuario} ({usuario.correo}) ha sido cancelada porque el estudiante fue eliminado del sistema.',
                        'sistema@example.com',
                        [asesoria.docente.correo],
                        fail_silently=True,
                    )
                except Exception as e:
                    print(f"Error al enviar correo de notificación: {e}")
            asesorias.delete()
        
        usuario.delete()
        return JsonResponse({"mensaje": "Usuario eliminado exitosamente"})
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def obtener_perfil_usuario(request):
    """Obtiene los detalles del perfil del usuario autenticado"""
    if request.method == "GET":
        if not request.user.is_authenticated:
            return JsonResponse({"mensaje": "No autorizado"}, status=401)
        
        try:
            usuario = request.user
            # Convertir last_login a la zona horaria local
            last_login = usuario.last_login
            if last_login:
                last_login = timezone.localtime(last_login).strftime("%Y-%m-%d %H:%M:%S")
            
            data = {
                "nombre_usuario": usuario.nombre_usuario,  # Incluimos el nombre del usuario
                "correo": usuario.correo,
                "rol": usuario.rol,
                "last_login": last_login,
                "is_active": usuario.is_active,
            }
            
            # Agregar campos adicionales específicos según el rol
            if usuario.rol == "Estudiante":
                data["semestre"] = usuario.semestre
                data["grupo"] = usuario.grupo
            elif usuario.rol == "Director":
                data["subtipo_director"] = usuario.subtipo_director
            elif usuario.rol == "Docente":
                data["subtipo_docente"] = usuario.subtipo_docente
            
            return JsonResponse(data)
        
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def cambiar_contraseña(request):
    """Permite al usuario cambiar su contraseña actual"""
    if request.method == "POST":
        if not request.user.is_authenticated:
            return JsonResponse({"mensaje": "No autorizado"}, status=401)
        
        try:
            data = json.loads(request.body)
            contraseña_actual = data.get("contraseña_actual")
            nueva_contraseña = data.get("nueva_contraseña")
            
            # Validar que se envíen los campos requeridos
            if not contraseña_actual or not nueva_contraseña:
                return JsonResponse({"mensaje": "La contraseña actual y la nueva contraseña son requeridas"}, status=400)
            
            # Verificar que la contraseña actual sea correcta
            usuario = request.user
            if not usuario.check_password(contraseña_actual):
                return JsonResponse({"mensaje": "La contraseña actual es incorrecta"}, status=400)
            
            # Validar complejidad de la nueva contraseña
            if len(nueva_contraseña) < 8:
                return JsonResponse({"mensaje": "La nueva contraseña debe tener al menos 8 caracteres"}, status=400)
            
            # Verificar que la nueva contraseña contenga al menos una letra mayúscula,
            # una minúscula y un número
            if not any(c.isupper() for c in nueva_contraseña):
                return JsonResponse({"mensaje": "La nueva contraseña debe contener al menos una letra mayúscula"}, status=400)
            
            if not any(c.islower() for c in nueva_contraseña):
                return JsonResponse({"mensaje": "La nueva contraseña debe contener al menos una letra minúscula"}, status=400)
                
            if not any(c.isdigit() for c in nueva_contraseña):
                return JsonResponse({"mensaje": "La nueva contraseña debe contener al menos un número"}, status=400)
            
            # Cambiar la contraseña
            usuario.set_password(nueva_contraseña)
            usuario.save()
            
            # Actualizar la sesión para que el usuario no tenga que iniciar sesión nuevamente
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, usuario)
            
            return JsonResponse({"mensaje": "Contraseña actualizada exitosamente"}, status=200)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)


@csrf_exempt
def crear_estudiante(request):
    """Crea un nuevo usuario con rol Estudiante y envía la contraseña por correo"""
    if request.method == "POST":
        if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
       
        try:
            data = json.loads(request.body)
           
            # Validar que se envíen los campos requeridos
            if "correo" not in data:
                return JsonResponse({"mensaje": "El correo es requerido"}, status=400)
            
            # Validar que se envíe el nombre (campo obligatorio)
            if "nombre_usuario" not in data:
                return JsonResponse({"mensaje": "El nombre es requerido"}, status=400)
           
            # Verificar si ya existe un usuario con el mismo correo
            if Usuario.objects.filter(correo=data["correo"]).exists():
                return JsonResponse({"mensaje": "Ya existe un usuario con este correo"}, status=400)
           
            # Generar contraseña aleatoria
            contraseña = generar_contraseña()
           
            # Crear el nuevo usuario (siempre como Estudiante)
            nuevo_estudiante = Usuario(
                nombre_usuario=data["nombre_usuario"],  # Incluimos el nombre del estudiante
                correo=data["correo"],
                rol="Estudiante",
                is_active=data.get("is_active", True),
                is_staff=False,
                is_superuser=False,
                semestre=data.get("semestre", ""),  # Agregar campo semestre
                grupo=data.get("grupo", "")         # Agregar campo grupo
            )
            nuevo_estudiante.set_password(contraseña)
            nuevo_estudiante.save()
           
            # Enviar correo con la contraseña
            send_mail(
                "Acceso al Sistema de Asesorías CESMAG",
                f"Hola {nuevo_estudiante.nombre_usuario},\n\nTú contraseña de acceso es: {contraseña}\n\n" +
                f"Y tu rol es: Estudiante." +
                f"\nSemestre: {nuevo_estudiante.semestre or 'No asignado'}" +
                f"\nGrupo: {nuevo_estudiante.grupo or 'No asignado'}",
                "lizcanoagudelo2@gmail.com",
                [data["correo"]],
                fail_silently=False,
            )
           
            return JsonResponse({
                "mensaje": "Estudiante creado exitosamente. Contraseña enviada por correo.",
                "id": nuevo_estudiante.id
            }, status=201)
           
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
   
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def listar_estudiantes(request):
    """Obtiene todos los usuarios con rol Estudiante"""
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            # Filtrar solo usuarios con rol Estudiante
            estudiantes = Usuario.objects.filter(rol="Estudiante")
            data = []
            
            for estudiante in estudiantes:
                data.append({
                    "id": estudiante.id,
                    "nombre_usuario": estudiante.nombre_usuario,  # Incluimos el nombre del estudiante
                    "correo": estudiante.correo,
                    "rol": estudiante.rol,
                    "semestre": estudiante.semestre,  # Agregar campo semestre
                    "grupo": estudiante.grupo,        # Agregar campo grupo
                    "is_active": estudiante.is_active
                })
            
            return JsonResponse(data, safe=False)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def detalle_estudiante(request, estudiante_id):
    """Obtiene, actualiza o elimina un usuario con rol Estudiante"""
    if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
        return JsonResponse({"mensaje": "No autorizado"}, status=403)
    
    try:
        # Verificar que exista y sea un Estudiante
        estudiante = Usuario.objects.get(id=estudiante_id, rol="Estudiante")
    except ObjectDoesNotExist:
        return JsonResponse({"mensaje": "Estudiante no encontrado"}, status=404)
    
    # GET: Obtener detalle de un estudiante
    if request.method == "GET":
        data = {
            "id": estudiante.id,
            "nombre_usuario": estudiante.nombre_usuario,  # Incluimos el nombre del estudiante
            "correo": estudiante.correo,
            "rol": estudiante.rol,
            "semestre": estudiante.semestre,  # Agregar campo semestre
            "grupo": estudiante.grupo,        # Agregar campo grupo
            "is_active": estudiante.is_active
        }
        return JsonResponse(data)
    
    # PUT: Actualizar un estudiante
    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            
            # Actualizar datos del estudiante (no se permite cambiar el rol)
            if "nombre_usuario" in data:
                estudiante.nombre_usuario = data["nombre_usuario"]
                
            if "correo" in data:
                # Verificar si el nuevo correo ya existe y no es del usuario actual
                if Usuario.objects.filter(correo=data["correo"]).exclude(id=estudiante_id).exists():
                    return JsonResponse({"mensaje": "El correo ya está en uso por otro usuario"}, status=400)
                estudiante.correo = data["correo"]
            
            # Ignorar cambios de rol si se intenta
            if "rol" in data and data["rol"] != "Estudiante":
                return JsonResponse({"mensaje": "No se permite cambiar el rol de un Estudiante"}, status=400)
            
            # Actualizar semestre y grupo
            if "semestre" in data:
                estudiante.semestre = data["semestre"]
                
            if "grupo" in data:
                estudiante.grupo = data["grupo"]
            
            if "is_active" in data:
                estudiante.is_active = data["is_active"]
            
            if "password" in data and data["password"]:
                estudiante.set_password(data["password"])
            
            estudiante.save()
            return JsonResponse({"mensaje": "Estudiante actualizado exitosamente"})
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # DELETE: Eliminar un estudiante
    elif request.method == "DELETE":
        estudiante.delete()
        return JsonResponse({"mensaje": "Estudiante eliminado exitosamente"})
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def crear_docente(request):
    """Crea un nuevo usuario con rol Docente y envía la contraseña por correo"""
    if request.method == "POST":
        if not request.user.is_authenticated or request.user.rol != "Director":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            data = json.loads(request.body)
            
            # Validar que se envíen los campos requeridos
            if "correo" not in data or "subtipo_docente" not in data:
                return JsonResponse({"mensaje": "El correo y subtipo del docente son requeridos"}, status=400)
            
            # Validar que se envíe el nombre (campo obligatorio)
            if "nombre_usuario" not in data:
                return JsonResponse({"mensaje": "El nombre es requerido"}, status=400)
            
            # Verificar que el subtipo_docente corresponda al subtipo_director
            subtipo_director = request.user.subtipo_director
            subtipo_docente_permitido = None
            
            if subtipo_director == "Director de Ing.Sistemas":
                subtipo_docente_permitido = "Docente de Ing.Sistemas"
            elif subtipo_director == "Director de Ing.Electrónica":
                subtipo_docente_permitido = "Docente de Ing.Electrónica"
            elif subtipo_director == "Director de Humanidades":
                subtipo_docente_permitido = "Docente de Humanidades"
            elif subtipo_director == "Director de Idiomas":
                subtipo_docente_permitido = "Docente de Idiomas"
            elif subtipo_director == "Director de Ciencias Básicas":
                subtipo_docente_permitido = "Docente de Ciencias Básicas"
            
            if data["subtipo_docente"] != subtipo_docente_permitido:
                return JsonResponse({
                    "mensaje": f"Como {subtipo_director}, solo puede crear docentes de tipo {subtipo_docente_permitido}"
                }, status=400)
            
            # Verificar si ya existe un usuario con el mismo correo
            if Usuario.objects.filter(correo=data["correo"]).exists():
                return JsonResponse({"mensaje": "Ya existe un usuario con este correo"}, status=400)
            
            # Generar contraseña aleatoria
            contraseña = generar_contraseña()
            
            # Crear el nuevo usuario (siempre como Docente)
            nuevo_docente = Usuario(
                nombre_usuario=data["nombre_usuario"],  # Incluimos el nombre del docente
                correo=data["correo"],
                rol="Docente",
                subtipo_docente=data["subtipo_docente"],
                is_active=data.get("is_active", True),
                is_staff=False,
                is_superuser=False
            )
            nuevo_docente.set_password(contraseña)
            nuevo_docente.save()
            
            # Enviar correo con la contraseña
            send_mail(
                "Acceso al Sistema de Asesorías CESMAG",
                f"Hola {nuevo_docente.nombre_usuario},\n\nTú contraseña de acceso es: {contraseña}\n\n" +
                f"Y tu rol es: Docente ({data['subtipo_docente']}).",
                "lizcanoagudelo2@gmail.com",
                [data["correo"]],
                fail_silently=False,
            )
            
            return JsonResponse({
                "mensaje": "Docente creado exitosamente. Contraseña enviada por correo.",
                "id": nuevo_docente.id
            }, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)


@csrf_exempt
def listar_docentes(request):
    """Obtiene todos los usuarios con rol Docente según el subtipo del director"""
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol != "Director":
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        try:
            # Determinar qué subtipo de docente puede gestionar este director
            subtipo_director = request.user.subtipo_director
            subtipo_docente_permitido = None
            
            if subtipo_director == "Director de Ing.Sistemas":
                subtipo_docente_permitido = "Docente de Ing.Sistemas"
            elif subtipo_director == "Director de Ing.Electrónica":
                subtipo_docente_permitido = "Docente de Ing.Electrónica"
            elif subtipo_director == "Director de Humanidades":
                subtipo_docente_permitido = "Docente de Humanidades"
            elif subtipo_director == "Director de Idiomas":
                subtipo_docente_permitido = "Docente de Idiomas"
            elif subtipo_director == "Director de Ciencias Básicas":
                subtipo_docente_permitido = "Docente de Ciencias Básicas"
            
            # Filtrar solo usuarios con rol Docente y el subtipo correspondiente
            docentes = Usuario.objects.filter(
                rol="Docente", 
                subtipo_docente=subtipo_docente_permitido
            )
            data = []
            
            for docente in docentes:
                data.append({
                    "id": docente.id,
                    "nombre_usuario": docente.nombre_usuario,  # Añadido el campo nombre
                    "correo": docente.correo,
                    "rol": docente.rol,
                    "subtipo_docente": docente.subtipo_docente,
                    "is_active": docente.is_active
                })
            
            return JsonResponse(data, safe=False)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)


@csrf_exempt
def detalle_docente(request, docente_id):
    """Obtiene, actualiza o elimina un usuario con rol Docente"""
    if not request.user.is_authenticated or request.user.rol != "Director":
        return JsonResponse({"mensaje": "No autorizado"}, status=403)
    
    # Determinar qué subtipo de docente puede gestionar este director
    subtipo_director = request.user.subtipo_director
    subtipo_docente_permitido = None
    
    if subtipo_director == "Director de Ing.Sistemas":
        subtipo_docente_permitido = "Docente de Ing.Sistemas"
    elif subtipo_director == "Director de Ing.Electrónica":
        subtipo_docente_permitido = "Docente de Ing.Electrónica"
    elif subtipo_director == "Director de Humanidades":
        subtipo_docente_permitido = "Docente de Humanidades"
    elif subtipo_director == "Director de Idiomas":
        subtipo_docente_permitido = "Docente de Idiomas"
    elif subtipo_director == "Director de Ciencias Básicas":
        subtipo_docente_permitido = "Docente de Ciencias Básicas"
    
    try:
        # Verificar que exista y sea un Docente del subtipo permitido
        docente = Usuario.objects.get(
            id=docente_id, 
            rol="Docente",
            subtipo_docente=subtipo_docente_permitido
        )
    except ObjectDoesNotExist:
        return JsonResponse({"mensaje": "Docente no encontrado o no tiene permisos para gestionarlo"}, status=404)
    
    # GET: Obtener detalle de un docente
    if request.method == "GET":
        data = {
            "id": docente.id,
            "nombre_usuario": docente.nombre_usuario,  # Añadido el campo nombre
            "correo": docente.correo,
            "rol": docente.rol,
            "subtipo_docente": docente.subtipo_docente,
            "is_active": docente.is_active
        }
        return JsonResponse(data)
    
    # PUT: Actualizar un docente
    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            
            # Actualizar datos del docente (no se permite cambiar el rol ni el subtipo)
            if "nombre_usuario" in data:
                docente.nombre_usuario = data["nombre_usuario"]  # Añadida la actualización del nombre
                
            if "correo" in data:
                # Verificar si el nuevo correo ya existe y no es del usuario actual
                if Usuario.objects.filter(correo=data["correo"]).exclude(id=docente_id).exists():
                    return JsonResponse({"mensaje": "El correo ya está en uso por otro usuario"}, status=400)
                docente.correo = data["correo"]
            
            # Ignorar cambios de rol si se intenta
            if "rol" in data and data["rol"] != "Docente":
                return JsonResponse({"mensaje": "No se permite cambiar el rol de un Docente"}, status=400)
            
            # Ignorar cambios de subtipo si se intenta
            if "subtipo_docente" in data and data["subtipo_docente"] != subtipo_docente_permitido:
                return JsonResponse({"mensaje": f"No se permite cambiar el subtipo de un {subtipo_docente_permitido}"}, status=400)
            
            if "is_active" in data:
                docente.is_active = data["is_active"]
            
            if "password" in data and data["password"]:
                docente.set_password(data["password"])
            
            docente.save()
            return JsonResponse({"mensaje": "Docente actualizado exitosamente"})
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # DELETE: Eliminar un docente
    elif request.method == "DELETE":
        docente.delete()
        return JsonResponse({"mensaje": "Docente eliminado exitosamente"})
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def gestionar_asignatura(request):
    """Crea, actualiza, elimina o lista asignaturas con sus áreas y docentes asociados"""
    
    # Verificar autenticación y permisos
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Solo administradores, directores y docentes pueden gestionar asignaturas
    if request.user.rol not in ["Administrador", "Director", "Docente"]:
        return JsonResponse({"mensaje": "No tiene permisos para gestionar asignaturas"}, status=403)
    
    # POST: Crear una nueva asignatura
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            
            # Validar campos requeridos
            campos_requeridos = ["nombre", "area", "semestre"]
            for campo in campos_requeridos:
                if campo not in data:
                    return JsonResponse({"mensaje": f"El campo '{campo}' es requerido"}, status=400)
            
            # Validar que el semestre sea válido
            semestres_validos = [choice[0] for choice in Asignatura.SEMESTRES]
            if data["semestre"] not in semestres_validos:
                return JsonResponse({
                    "mensaje": f"Semestre no válido. Opciones válidas: {', '.join(semestres_validos)}"
                }, status=400)
            
            # Obtener o crear el área
            area_nombre = data["area"]
            
            # Si es director, buscar o crear área solo dentro de su subtipo
            if request.user.rol == "Director":
                try:
                    area = Area.objects.get(
                        nombre=area_nombre, 
                        subtipo_director=request.user.subtipo_director
                    )
                except ObjectDoesNotExist:
                    # Si el área no existe, crearla asociada al director actual
                    area = Area.objects.create(
                        nombre=area_nombre,
                        director_creador=request.user,
                        subtipo_director=request.user.subtipo_director
                    )
            else:
                # Para administradores, buscar área por nombre sin restricción
                try:
                    area = Area.objects.get(nombre=area_nombre)
                except ObjectDoesNotExist:
                    # Si es administrador, asociar el área a sí mismo
                    area = Area.objects.create(
                        nombre=area_nombre,
                        director_creador=request.user,
                        subtipo_director="Administración"  # Valor predeterminado para administradores
                    )
            
            # Verificar si ya existe una asignatura con el mismo nombre y área
            if Asignatura.objects.filter(nombre=data["nombre"], area=area).exists():
                return JsonResponse({
                    "mensaje": f"Ya existe una asignatura con el nombre '{data['nombre']}' en el área '{area_nombre}'"
                }, status=400)
            
            # Determinar si usar los docentes del área o asignar docentes específicos
            usar_docentes_area = data.get("usar_docentes_area", True)
            
            # Crear la asignatura
            nueva_asignatura = Asignatura(
                nombre=data["nombre"],
                area=area,
                semestre=data["semestre"],
                director_creador=request.user,
                subtipo_director=request.user.subtipo_director if request.user.rol == "Director" else "Administración",
                usar_docentes_area=usar_docentes_area
            )
            nueva_asignatura.save()
            
            # Agregar docentes específicos sólo si no se usan los docentes del área
            if not usar_docentes_area and "docentes" in data and isinstance(data["docentes"], list):
                for correo_docente in data["docentes"]:
                    try:
                        docente = Usuario.objects.get(correo=correo_docente, rol="Docente")
                        nueva_asignatura.docentes.add(docente)
                    except ObjectDoesNotExist:
                        # No agregamos el docente si no existe o no es un docente
                        pass
            
            # Obtener docentes para la respuesta
            if usar_docentes_area:
                docentes = list(area.docentes.values_list('correo', flat=True))
            else:
                docentes = list(nueva_asignatura.docentes.values_list('correo', flat=True))
            
            return JsonResponse({
                "mensaje": "Asignatura creada exitosamente",
                "id": nueva_asignatura.id,
                "nombre": nueva_asignatura.nombre,
                "area": nueva_asignatura.area.nombre,
                "semestre": nueva_asignatura.semestre,
                "semestre_display": nueva_asignatura.get_semestre_display(),
                "usar_docentes_area": usar_docentes_area,
                "docentes": docentes,
                "director_creador": nueva_asignatura.director_creador.correo,
                "subtipo_director": nueva_asignatura.subtipo_director
            }, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # GET: Listar todas las asignaturas o filtrar por parámetros
    elif request.method == "GET":
        
        try:
            
            # Obtener parámetros de filtrado
            area_id = request.GET.get('area')
            semestre = request.GET.get('semestre')
            docente_id = request.GET.get('docente')
            
            # Iniciar consulta base
            # Si es director, filtrar solo sus asignaturas
            if request.user.rol == "Director" and request.user.subtipo_director:
                asignaturas = Asignatura.objects.filter(subtipo_director=request.user.subtipo_director)
            else:
                asignaturas = Asignatura.objects.all()
            
            # Aplicar filtros si se proporcionan
            if area_id:
                try:
                    asignaturas = asignaturas.filter(area_id=int(area_id))
                except ValueError:
                    return JsonResponse({"mensaje": "ID de área inválido"}, status=400)
            
            if semestre:
                asignaturas = asignaturas.filter(semestre=semestre)
            
            if docente_id:
                try:
                    # Este filtrado debe considerar tanto los docentes propios como los del área
                    docente_id = int(docente_id)
                    # Crear una lista para almacenar las asignaturas que cumplen el criterio
                    asignaturas_filtradas = []
                    
                    for asignatura in asignaturas:
                        if asignatura.usar_docentes_area:
                            # Si la asignatura usa docentes del área, verificar si el docente pertenece al área
                            if asignatura.area.docentes.filter(id=docente_id).exists():
                                asignaturas_filtradas.append(asignatura.id)
                        else:
                            # Si usa docentes propios, verificar si el docente está en la asignatura
                            if asignatura.docentes.filter(id=docente_id).exists():
                                asignaturas_filtradas.append(asignatura.id)
                    
                    # Filtrar por los IDs encontrados
                    asignaturas = asignaturas.filter(id__in=asignaturas_filtradas)
                except ValueError:
                    return JsonResponse({"mensaje": "ID de docente inválido"}, status=400)
            
            # Formatear la respuesta
            resultado = []
            for asignatura in asignaturas:
                # Determinar el subtipo_docente correspondiente según el subtipo_director
                subtipo_docente_correspondiente = None
                if asignatura.subtipo_director == "Director de Ing.Sistemas":
                    subtipo_docente_correspondiente = "Docente de Ing.Sistemas"
                elif asignatura.subtipo_director == "Director de Ing.Electrónica":
                    subtipo_docente_correspondiente = "Docente de Ing.Electrónica"
                elif asignatura.subtipo_director == "Director de Ciencias Básicas":
                    subtipo_docente_correspondiente = "Docente de Ciencias Básicas"
                elif asignatura.subtipo_director == "Director de Humanidades":
                    subtipo_docente_correspondiente = "Docente de Humanidades"
                elif asignatura.subtipo_director == "Director de Idiomas":
                    subtipo_docente_correspondiente = "Docente de Idiomas"
                
                # Obtener los docentes correctos según el valor de usar_docentes_area y filtrar por subtipo correspondiente
                if asignatura.usar_docentes_area:
                    if subtipo_docente_correspondiente:
                        docentes_info = list(asignatura.area.docentes.filter(
                            rol="Docente", 
                            subtipo_docente=subtipo_docente_correspondiente
                        ).values('id', 'correo', 'subtipo_docente'))
                    else:
                        # Si es un administrador u otro caso, no filtramos por subtipo
                        docentes_info = list(asignatura.area.docentes.filter(
                            rol="Docente"
                        ).values('id', 'correo', 'subtipo_docente'))
                else:
                    if subtipo_docente_correspondiente:
                        docentes_info = list(asignatura.docentes.filter(
                            rol="Docente", 
                            subtipo_docente=subtipo_docente_correspondiente
                        ).values('id', 'correo', 'subtipo_docente'))
                    else:
                        # Si es un administrador u otro caso, no filtramos por subtipo
                        docentes_info = list(asignatura.docentes.filter(
                            rol="Docente"
                        ).values('id', 'correo', 'subtipo_docente'))
                
                resultado.append({
                    "id": asignatura.id,
                    "nombre": asignatura.nombre,
                    "area": {
                        "id": asignatura.area.id,
                        "nombre": asignatura.area.nombre
                    },
                    "semestre": asignatura.semestre,
                    "semestre_display": asignatura.get_semestre_display(),
                    "usar_docentes_area": asignatura.usar_docentes_area,
                    "docentes": docentes_info,
                    "director_creador": asignatura.director_creador.correo,
                    "subtipo_director": asignatura.subtipo_director
                })
            
            return JsonResponse({"asignaturas": resultado})
            
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def detalle_asignatura(request, asignatura_id):
    """Obtiene, actualiza o elimina una asignatura específica"""
    # Verificar autenticación y permisos
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Solo administradores, directores y docentes pueden acceder al detalle
    if request.user.rol not in ["Administrador", "Director", "Docente"]:
        return JsonResponse({"mensaje": "No tiene permisos para gestionar asignaturas"}, status=403)
    
    try:
        asignatura = Asignatura.objects.get(id=asignatura_id)
        
        # Si es director, verificar que la asignatura pertenezca a su subtipo
        if (request.user.rol == "Director" and 
            request.user.subtipo_director and 
            asignatura.subtipo_director != request.user.subtipo_director):
            return JsonResponse({"mensaje": "No tiene acceso a esta asignatura"}, status=403)
            
    except ObjectDoesNotExist:
        return JsonResponse({"mensaje": "Asignatura no encontrada"}, status=404)
    
    # GET: Obtener detalle de una asignatura
    if request.method == "GET":
        # Obtener los docentes correctos según el valor de usar_docentes_area
        if asignatura.usar_docentes_area:
            # Filtrar para incluir solo usuarios que actualmente son docentes
            docentes_info = list(asignatura.area.docentes.filter(rol="Docente").values('id', 'correo', 'subtipo_docente'))
        else:
            # Filtrar para incluir solo usuarios que actualmente son docentes
            docentes_info = list(asignatura.docentes.filter(rol="Docente").values('id', 'correo', 'subtipo_docente'))
            
        data = {
            "id": asignatura.id,
            "nombre": asignatura.nombre,
            "area": {
                "id": asignatura.area.id,
                "nombre": asignatura.area.nombre
            },
            "semestre": asignatura.semestre,
            "semestre_display": asignatura.get_semestre_display(),
            "usar_docentes_area": asignatura.usar_docentes_area,
            "docentes": docentes_info,
            "director_creador": asignatura.director_creador.correo,
            "subtipo_director": asignatura.subtipo_director
        }
        return JsonResponse(data)
    
    # PUT: Actualizar una asignatura
    elif request.method == "PUT":
        # Verificar si el usuario puede editar (admin o cualquier director del mismo subtipo)
        if not (request.user.rol == "Administrador" or 
                (request.user.rol == "Director" and 
                 request.user.subtipo_director == asignatura.subtipo_director)):
            return JsonResponse({"mensaje": "No tiene permisos para editar esta asignatura"}, status=403)
        
        try:
            data = json.loads(request.body)
            
            # Actualizar nombre si se proporciona
            if "nombre" in data:
                asignatura.nombre = data["nombre"]
            
            # Actualizar área si se proporciona
            if "area" in data:
                area_nombre = data["area"]
                
                # Si es director, buscar área dentro de su subtipo
                if request.user.rol == "Director":
                    try:
                        area = Area.objects.get(
                            nombre=area_nombre, 
                            subtipo_director=request.user.subtipo_director
                        )
                    except ObjectDoesNotExist:
                        # Si el área no existe, crearla asociada al director actual
                        area = Area.objects.create(
                            nombre=area_nombre,
                            director_creador=request.user,
                            subtipo_director=request.user.subtipo_director
                        )
                else:
                    # Para administradores, buscar área por nombre
                    try:
                        area = Area.objects.get(nombre=area_nombre)
                    except ObjectDoesNotExist:
                        # Si es administrador, crear área asociada a sí mismo
                        area = Area.objects.create(
                            nombre=area_nombre,
                            director_creador=request.user,
                            subtipo_director="Administración"
                        )
                        
                asignatura.area = area
            
            # Actualizar semestre si se proporciona
            if "semestre" in data:
                # Validar que el semestre sea válido
                semestres_validos = [choice[0] for choice in Asignatura.SEMESTRES]
                if data["semestre"] not in semestres_validos:
                    return JsonResponse({
                        "mensaje": f"Semestre no válido. Opciones válidas: {', '.join(semestres_validos)}"
                    }, status=400)
                asignatura.semestre = data["semestre"]
            
            # Actualizar el campo usar_docentes_area si se proporciona
            if "usar_docentes_area" in data:
                asignatura.usar_docentes_area = bool(data["usar_docentes_area"])
            
            # Actualizar docentes específicos sólo si no se usan los docentes del área
            if not asignatura.usar_docentes_area and "docentes" in data and isinstance(data["docentes"], list):
                # Eliminar docentes actuales
                asignatura.docentes.clear()
                # Agregar nuevos docentes
                for correo_docente in data["docentes"]:
                    try:
                        docente = Usuario.objects.get(correo=correo_docente, rol="Docente")
                        asignatura.docentes.add(docente)
                    except ObjectDoesNotExist:
                        # No agregamos el docente si no existe o no es un docente
                        pass
            
            asignatura.save()
            return JsonResponse({"mensaje": "Asignatura actualizada exitosamente"})
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # DELETE: Eliminar una asignatura
    elif request.method == "DELETE":
        # Solo administradores o directores del mismo subtipo pueden eliminar
        if not (request.user.rol == "Administrador" or 
                (request.user.rol == "Director" and 
                 request.user.subtipo_director == asignatura.subtipo_director)):
            return JsonResponse({"mensaje": "No tiene permisos para eliminar esta asignatura"}, status=403)
        
        asignatura.delete()
        return JsonResponse({"mensaje": "Asignatura eliminada exitosamente"})
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

# Los métodos listar_areas y detalle_area permanecen casi igual, pero necesitamos
# añadir algunos cambios menores para mostrar las asignaturas correctamente

@csrf_exempt
def listar_areas(request):
    """Lista todas las áreas disponibles para las asignaturas"""
    
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Verificar que el método sea GET
    if request.method != "GET":
        return JsonResponse({"mensaje": "Método no permitido"}, status=405)
    
    try:
        # Si es director, mostrar solo áreas de su subtipo
        if request.user.rol == "Director" and request.user.subtipo_director:
            areas = Area.objects.filter(subtipo_director=request.user.subtipo_director).order_by('nombre')
        else:
            # Para administradores y otros roles, mostrar todas
            areas = Area.objects.all().order_by('nombre')
        
        # Formatear la respuesta
        resultado = []
        for area in areas:
            # Contar cuántas asignaturas tiene cada área
            num_asignaturas = area.asignaturas.count()
            
            # Determinar el subtipo_docente correspondiente según el subtipo_director
            subtipo_docente_correspondiente = None
            if area.subtipo_director == "Director de Ing.Sistemas":
                subtipo_docente_correspondiente = "Docente de Ing.Sistemas"    
            elif area.subtipo_director == "Director de Ing.Electrónica":
                subtipo_docente_correspondiente = "Docente de Ing.Electrónica"    
            elif area.subtipo_director == "Director de Ciencias Básicas":
                subtipo_docente_correspondiente = "Docente de Ciencias Básicas"
            elif area.subtipo_director == "Director de Humanidades":
                subtipo_docente_correspondiente = "Docente de Humanidades"
            elif area.subtipo_director == "Director de Idiomas":
                subtipo_docente_correspondiente = "Docente de Idiomas"
            
            # Filtrar para incluir solo usuarios que actualmente son docentes y del subtipo correspondiente
            if subtipo_docente_correspondiente:
                docentes_actuales = area.docentes.filter(
                    rol="Docente", 
                    subtipo_docente=subtipo_docente_correspondiente
                )
            else:
                # Si es administración u otro caso, no filtrar por subtipo
                docentes_actuales = area.docentes.filter(rol="Docente")
            
            resultado.append({
                "id": area.id,
                "nombre": area.nombre,
                "num_asignaturas": num_asignaturas,
                "director_creador": area.director_creador.correo,
                "subtipo_director": area.subtipo_director,
                # Incluir solo los docentes con rol actual "Docente" y del subtipo correspondiente
                "docentes": list(docentes_actuales.values('id', 'correo', 'subtipo_docente')),
                # Opcionalmente, puedes incluir las primeras N asignaturas del área
                "asignaturas_ejemplo": list(area.asignaturas.values('id', 'nombre')[:3])
            })
        
        return JsonResponse({"areas": resultado})
        
    except Exception as e:
        return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
@csrf_exempt
def detalle_area(request, area_id):
    """Obtiene, actualiza o elimina un área específica"""
    # Verificar autenticación y permisos
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Solo administradores y directores pueden gestionar áreas
    if request.user.rol not in ["Administrador", "Director"]:
        return JsonResponse({"mensaje": "No tiene permisos para gestionar áreas"}, status=403)
    
    try:
        area = Area.objects.get(id=area_id)
        
        # Si es director, verificar que el área pertenece a su subtipo
        if (request.user.rol == "Director" and 
            request.user.subtipo_director and 
            area.subtipo_director != request.user.subtipo_director):
            return JsonResponse({"mensaje": "No tiene acceso a esta área"}, status=403)
            
    except ObjectDoesNotExist:
        return JsonResponse({"mensaje": "Área no encontrada"}, status=404)
    
    # GET: Obtener detalle de un área
    if request.method == "GET":
        # Contar asignaturas relacionadas
        num_asignaturas = area.asignaturas.count()
        
        data = {
            "id": area.id,
            "nombre": area.nombre,
            "director_creador": area.director_creador.correo,
            "subtipo_director": area.subtipo_director,
            "num_asignaturas": num_asignaturas,
            "docentes": list(area.docentes.values('id', 'correo', 'subtipo_docente')),
            "asignaturas": list(area.asignaturas.values('id', 'nombre', 'semestre', 'usar_docentes_area'))
        }
        return JsonResponse(data)
    
    # PUT: Actualizar un área
    elif request.method == "PUT":
        # Verificar si el usuario puede editar (admin o cualquier director del mismo subtipo)
        if not (request.user.rol == "Administrador" or 
                (request.user.rol == "Director" and 
                 request.user.subtipo_director == area.subtipo_director)):
            return JsonResponse({"mensaje": "No tiene permisos para editar esta área"}, status=403)
            
        try:
            data = json.loads(request.body)
            
            # Actualizar nombre si se proporciona
            if "nombre" in data:
                # Verificar si ya existe un área con ese nombre (en cualquier departamento)
                if Area.objects.filter(nombre=data["nombre"]).exclude(id=area_id).exists():
                    return JsonResponse({"mensaje": "Ya existe un área con ese nombre"}, status=400)
                
                area.nombre = data["nombre"]
            
            # Actualizar docentes si se proporcionan
            docentes_no_encontrados = []
            docentes_encontrados = []
            
            if "docentes" in data and isinstance(data["docentes"], list):
                # Validar que haya docentes asignados
                if len(data["docentes"]) == 0:
                    return JsonResponse({"mensaje": "Se requiere asignar al menos un docente al área"}, status=400)
                
                # Eliminar docentes actuales
                area.docentes.clear()
                
                # Agregar nuevos docentes
                for correo_docente in data["docentes"]:
                    try:
                        # Primero hacemos una búsqueda menos restrictiva, solo por correo
                        try:
                            docente = Usuario.objects.get(correo__iexact=correo_docente)
                            
                            # Verificamos que sea docente
                            if docente.rol != "Docente":
                                docentes_no_encontrados.append(f"{correo_docente} (no es un docente)")
                                continue
                                
                            # Si es director, verificamos que sea del mismo subtipo
                            if request.user.rol == "Director":
                                # Extrae el departamento (parte después de "Docente de" o "Director de")
                                departamento_docente = docente.subtipo_docente.replace("Docente de ", "").strip() 
                                departamento_director = request.user.subtipo_director.replace("Director de ", "").strip()
                                departamento_docente2 = docente.subtipo_docente.replace("Docente del ", "").strip() 
                                departamento_director2 = request.user.subtipo_director.replace("Director del ", "").strip()

                                # Compara solo los departamentos
                                if departamento_docente != departamento_director and departamento_docente2 != departamento_director2:
                                    docentes_no_encontrados.append(f"{correo_docente} (no es de tu departamento)")
                                    continue
                                
                            # Si pasó todas las validaciones, lo agregamos
                            area.docentes.add(docente)
                            docentes_encontrados.append(correo_docente)
                            
                        except ObjectDoesNotExist:
                            # No se encontró el correo
                            docentes_no_encontrados.append(correo_docente)
                            
                    except Exception as e:
                        # Error inesperado
                        docentes_no_encontrados.append(f"{correo_docente} (error: {str(e)})")
                
                # Verificar que al menos se haya agregado un docente
                if not docentes_encontrados:
                    return JsonResponse({
                        "mensaje": "No se pudo actualizar el área porque no se encontraron docentes válidos",
                        "docentes_no_encontrados": docentes_no_encontrados
                    }, status=400)
            
            area.save()
            
            response_data = {
                "mensaje": "Área actualizada exitosamente",
                "id": area.id,
                "nombre": area.nombre,
                "docentes": list(area.docentes.values('id', 'correo', 'subtipo_docente')),
                "docentes_encontrados": docentes_encontrados
            }
            
            # Si hubo docentes que no se pudieron agregar, notificarlo
            if docentes_no_encontrados:
                response_data["advertencia"] = f"No se encontraron los siguientes docentes: {', '.join(docentes_no_encontrados)}"
            
            return JsonResponse(response_data)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # DELETE: Eliminar un área
    elif request.method == "DELETE":
        # Solo administradores o directores del mismo subtipo pueden eliminar
        if not (request.user.rol == "Administrador" or 
                (request.user.rol == "Director" and 
                 request.user.subtipo_director == area.subtipo_director)):
            return JsonResponse({"mensaje": "No tiene permisos para eliminar esta área"}, status=403)
            
        # Verificar si el área tiene asignaturas asociadas
        if area.asignaturas.exists():
            return JsonResponse({
                "mensaje": "No se puede eliminar el área porque tiene asignaturas asociadas",
                "num_asignaturas": area.asignaturas.count(),
                "asignaturas": list(area.asignaturas.values('id', 'nombre')[:5])  # Mostrar hasta 5 ejemplos
            }, status=400)
        
        area.delete()
        return JsonResponse({"mensaje": "Área eliminada exitosamente"})
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def listar_docentes_por_subtipo(request):
    """Lista los docentes filtrados por el subtipo del director actual"""
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Verificar que el método sea GET
    if request.method != "GET":
        return JsonResponse({"mensaje": "Método no permitido"}, status=405)
    
    try:
        # Si es director, mostrar solo docentes de su mismo subtipo
        if request.user.rol == "Director" and request.user.subtipo_director:
            # Mapeo de subtipo_director a subtipo_docente
            subtipo_docente_correspondiente = request.user.subtipo_director
            
            # Si hay un mapeo específico (ej. "Director del Programa" -> "Docente del Programa")
            if request.user.subtipo_director == "Director de Ing.Sistemas":
                subtipo_docente_correspondiente = "Docente de Ing.Sistemas"  
            elif request.user.subtipo_director == "Director de Ing.Electrónica":
                subtipo_docente_correspondiente = "Docente de Ing.Electrónica"    
            elif request.user.subtipo_director == "Director de Humanidades":
                subtipo_docente_correspondiente = "Docente de Humanidades"
            elif request.user.subtipo_director == "Director de Ciencias Básicas":
                subtipo_docente_correspondiente = "Docente de Ciencias Básicas"
            elif request.user.subtipo_director == "Director de Idiomas":
                subtipo_docente_correspondiente = "Docente de Idiomas"
            # Añadir más mapeos según sea necesario
            
            docentes = Usuario.objects.filter(
                rol="Docente", 
                subtipo_docente=subtipo_docente_correspondiente
            ).order_by('nombre')
        else:
            # Para administradores y otros roles, mostrar todos los docentes
            docentes = Usuario.objects.filter(rol="Docente").order_by('nombre')
        
        resultado = list(docentes.values('id', 'nombre', 'correo', 'subtipo_docente'))
        return JsonResponse(resultado, safe=False)
        
    except Exception as e:
        return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)

# Esta función adicional puede ayudar a listar docentes por área
@csrf_exempt
def listar_docentes_por_area(request, area_id):
    """Lista los docentes asociados a un área específica"""
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Verificar que el método sea GET
    if request.method != "GET":
        return JsonResponse({"mensaje": "Método no permitido"}, status=405)
    
    try:
        # Obtener el área
        try:
            area = Area.objects.get(id=area_id)
            
            # Si es director, verificar que el área pertenece a su subtipo
            if (request.user.rol == "Director" and 
                request.user.subtipo_director and 
                area.subtipo_director != request.user.subtipo_director):
                return JsonResponse({"mensaje": "No tiene acceso a esta área"}, status=403)
                
        except ObjectDoesNotExist:
            return JsonResponse({"mensaje": "Área no encontrada"}, status=404)
        
        # Obtener los docentes del área
        docentes = area.docentes.all().order_by('nombre')
        resultado = list(docentes.values('id', 'nombre', 'correo', 'subtipo_docente'))
        
        return JsonResponse({
            "area": area.nombre,
            "docentes": resultado
        })
        
    except Exception as e:
        return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
@csrf_exempt
def gestionar_area(request):
    """Crea o lista áreas con sus docentes asociados"""
    
    # Verificar autenticación y permisos
    if not request.user.is_authenticated:
        return JsonResponse({"mensaje": "No autorizado"}, status=401)
    
    # Solo administradores y directores pueden gestionar áreas
    if request.user.rol not in ["Administrador", "Director"]:
        return JsonResponse({"mensaje": "No tiene permisos para gestionar áreas"}, status=403)
    
    # POST: Crear una nueva área
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            
            # Validar campos requeridos
            if "nombre" not in data:
                return JsonResponse({"mensaje": "El nombre del área es requerido"}, status=400)
                
            # Validar que haya docentes asignados
            if "docentes" not in data or not isinstance(data["docentes"], list) or len(data["docentes"]) == 0:
                return JsonResponse({"mensaje": "Se requiere asignar al menos un docente al área"}, status=400)
            
            # Verificar si ya existe un área con ese nombre (en cualquier departamento)
            if Area.objects.filter(nombre=data["nombre"]).exists():
                return JsonResponse({"mensaje": "Ya existe un área con ese nombre"}, status=400)
            
            # Crear el área
            nueva_area = Area(
                nombre=data["nombre"],
                director_creador=request.user,
                subtipo_director=request.user.subtipo_director if request.user.rol == "Director" else "Administración"
            )
            nueva_area.save()
            
            # Agregar docentes si se proporcionan
            docentes_no_encontrados = []
            docentes_encontrados = []
            
            for correo_docente in data["docentes"]:
                try:
                    # Primero hacemos una búsqueda menos restrictiva, solo por correo
                    try:
                        docente = Usuario.objects.get(correo__iexact=correo_docente)
                        
                        # Verificamos que sea docente
                        if docente.rol != "Docente":
                            docentes_no_encontrados.append(f"{correo_docente} (no es un docente)")
                            continue
                            
                        # Si es director, verificamos que sea del mismo subtipo
                        if request.user.rol == "Director":
                            # Extrae el departamento (parte después de "Docente de" o "Director de")
                            departamento_docente = docente.subtipo_docente.replace("Docente de ", "").strip() 
                            departamento_director = request.user.subtipo_director.replace("Director de ", "").strip()
                            departamento_docente2 = docente.subtipo_docente.replace("Docente del ", "").strip() 
                            departamento_director2 = request.user.subtipo_director.replace("Director del ", "").strip()

                            # Compara solo los departamentos
                            if departamento_docente != departamento_director and departamento_docente2 != departamento_director2:
                                docentes_no_encontrados.append(f"{correo_docente} (no es de tu departamento)")
                                continue
                            
                        # Si pasó todas las validaciones, lo agregamos
                        nueva_area.docentes.add(docente)
                        docentes_encontrados.append(correo_docente)
                        
                    except ObjectDoesNotExist:
                        # No se encontró el correo
                        docentes_no_encontrados.append(correo_docente)
                        
                except Exception as e:
                    # Error inesperado
                    docentes_no_encontrados.append(f"{correo_docente} (error: {str(e)})")
            
            # Verificar que al menos se haya agregado un docente
            if not docentes_encontrados:
                # Si no se agregó ningún docente, eliminamos el área creada
                nueva_area.delete()
                return JsonResponse({
                    "mensaje": "No se pudo crear el área porque no se encontraron docentes válidos",
                    "docentes_no_encontrados": docentes_no_encontrados
                }, status=400)
            
            response_data = {
                "mensaje": "Área creada exitosamente",
                "id": nueva_area.id,
                "nombre": nueva_area.nombre,
                "director_creador": nueva_area.director_creador.correo,
                "subtipo_director": nueva_area.subtipo_director,
                "docentes": list(nueva_area.docentes.values('id', 'correo', 'subtipo_docente')),
                "docentes_encontrados": docentes_encontrados
            }
            
            # Si hubo docentes que no se pudieron agregar, notificarlo
            if docentes_no_encontrados:
                response_data["advertencia"] = f"No se encontraron los siguientes docentes: {', '.join(docentes_no_encontrados)}"
            
            return JsonResponse(response_data, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # GET: Listar todas las áreas
    elif request.method == "GET":
        try:
            # Si es director, mostrar solo áreas de su subtipo
            if request.user.rol == "Director" and request.user.subtipo_director:
                areas = Area.objects.filter(subtipo_director=request.user.subtipo_director).order_by('nombre')
            else:
                # Para administradores y otros roles, mostrar todas
                areas = Area.objects.all().order_by('nombre')
            
            # Formatear la respuesta
            resultado = []
            for area in areas:
                # Determinar el subtipo_docente correspondiente según el subtipo_director
                subtipo_docente_correspondiente = None
                if area.subtipo_director == "Director de Ing.Sistemas":
                    subtipo_docente_correspondiente = "Docente de Ing.Sistemas"    
                elif area.subtipo_director == "Director de Ing.Electrónica":
                    subtipo_docente_correspondiente = "Docente de Ing.Electrónica"
                elif area.subtipo_director == "Director de Ciencias Básicas":
                    subtipo_docente_correspondiente = "Docente de Ciencias Básicas"
                elif area.subtipo_director == "Director de Humanidades":
                    subtipo_docente_correspondiente = "Docente de Humanidades"
                elif area.subtipo_director == "Director de Idiomas":
                    subtipo_docente_correspondiente = "Docente de Idiomas"
                
                # Filtrar los docentes por el subtipo correspondiente si aplica
                if subtipo_docente_correspondiente:
                    docentes_filtrados = area.docentes.filter(
                        rol="Docente", 
                        subtipo_docente=subtipo_docente_correspondiente
                    )
                else:
                    # Si es un administrador u otro caso, no filtramos por subtipo
                    docentes_filtrados = area.docentes.filter(rol="Docente")
                
                resultado.append({
                    "id": area.id,
                    "nombre": area.nombre,
                    "director_creador": area.director_creador.correo,
                    "subtipo_director": area.subtipo_director,
                    "num_asignaturas": area.asignaturas.count(),
                    "docentes": list(docentes_filtrados.values('id', 'correo', 'subtipo_docente'))
                })
            
            return JsonResponse({"areas": resultado})
            
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

# In views.py, add this function
@csrf_exempt
def obtener_asignaturas_docente(request):
    """
    Obtiene las asignaturas asignadas a un docente, ya sea directamente
    o a través de las áreas a las que pertenece.
    """
    if request.method == "GET":
        try:
            # Obtener el usuario actual
            usuario = request.user
            
            if not usuario.is_authenticated:
                return JsonResponse({"error": "Usuario no autenticado"}, status=401)
            
            if usuario.rol != "Docente":
                return JsonResponse({"error": "Solo los docentes pueden acceder a esta función"}, status=403)
            
            # Asignaturas directamente asignadas al docente
            asignaturas_directas = Asignatura.objects.filter(docentes=usuario)
            
            # Asignaturas asignadas a través de áreas
            areas_docente = Area.objects.filter(docentes=usuario)
            asignaturas_areas = Asignatura.objects.filter(area__in=areas_docente, usar_docentes_area=True)
            
            # Combinar y eliminar duplicados
            todas_asignaturas = asignaturas_directas.union(asignaturas_areas)
            
            # Preparar la respuesta
            asignaturas_data = []
            for asignatura in todas_asignaturas:
                area_data = {
                    "id": asignatura.area.id,
                    "nombre": asignatura.area.nombre,
                    "subtipo_director": asignatura.area.subtipo_director
                }
                
                asignaturas_data.append({
                    "id": asignatura.id,
                    "nombre": asignatura.nombre,
                    "semestre": asignatura.get_semestre_display(),
                    "area": area_data,
                    "asignado_directamente": asignatura in asignaturas_directas
                })
            
            return JsonResponse({"asignaturas": asignaturas_data}, status=200)
            
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    
    return JsonResponse({"error": "Método no permitido"}, status=405)

@csrf_exempt
def listar_periodos(request):
    """Obtiene todos los periodos académicos"""
    if request.method == "GET":
        if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
            return JsonResponse({"mensaje": "No autorizado"}, status=403)
        
        periodos = Periodo.objects.all().order_by('-codigo')
        data = []
        
        # Determinar si hay un periodo activo y en curso
        fecha_actual = datetime.now().date()
        periodo_activo = Periodo.objects.filter(
            activo=True, 
            fecha_inicio__lte=fecha_actual, 
            fecha_fin__gte=fecha_actual
        ).first()
        
        for periodo in periodos:
            # Determinar si este periodo es editable
            es_editable = True
            if periodo_activo and periodo.id != periodo_activo.id:
                es_editable = False
                
            data.append({
                "id": periodo.id,
                "codigo": periodo.codigo,
                "fecha_inicio": periodo.fecha_inicio,
                "fecha_fin": periodo.fecha_fin,
                "activo": periodo.activo,
                "editable": es_editable
            })
        
        return JsonResponse(data, safe=False)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

@csrf_exempt
def crear_periodo(request):
    """Crea un nuevo periodo académico"""
    if request.method == "POST":
        if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
            return JsonResponse({"mensaje": "No autorizado"}, status=403)

        # Validar si ya existe un periodo que no ha finalizado
        fecha_actual = datetime.now().date()
        periodo_activo = Periodo.objects.filter(
            activo=True, 
            fecha_fin__gte=fecha_actual
        ).first()
        
        # Si hay un periodo activo y no ha finalizado, no permitir crear uno nuevo
        if periodo_activo:
            return JsonResponse({
                "mensaje": "No se puede crear un nuevo periodo mientras exista un periodo activo en curso",
                "periodo_vigente": {
                    "codigo": periodo_activo.codigo,
                    "fecha_inicio": periodo_activo.fecha_inicio,
                    "fecha_fin": periodo_activo.fecha_fin,
                    "activo": periodo_activo.activo
                }
            }, status=400)

        try:
            data = json.loads(request.body)

            # Validar que el periodo esté activo para crearlo
            if not data.get("activo", True):
                return JsonResponse({
                    "mensaje": "No se puede crear un periodo inactivo. El periodo debe estar activo al crearlo."
                }, status=400)

            # Obtener fechas
            fecha_inicio = data.get("fecha_inicio")
            fecha_fin = data.get("fecha_fin")

            if not fecha_inicio or not fecha_fin:
                return JsonResponse({"mensaje": "Las fechas de inicio y fin son obligatorias"}, status=400)

            # Convertir strings a objetos de fecha
            fecha_inicio_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_fin_obj = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

            # Obtener el año actual
            año_actual = datetime.now().year

            # Validar que ambas fechas pertenezcan al mismo año
            if fecha_inicio_obj.year != fecha_fin_obj.year:
                return JsonResponse({
                    "mensaje": "Las fechas de inicio y fin deben pertenecer al mismo año académico"
                }, status=400)

            # Validar que ambas fechas sean del año actual
            if fecha_inicio_obj.year != año_actual or fecha_fin_obj.year != año_actual:
                return JsonResponse({
                    "mensaje": "Las fechas de inicio y fin deben ser del año actual"
                }, status=400)

            # Validar que las fechas no sean pasadas
            if fecha_inicio_obj < fecha_actual or fecha_fin_obj < fecha_actual:
                return JsonResponse({
                    "mensaje": "Las fechas de inicio y fin no pueden ser anteriores a la fecha actual"
                }, status=400)

            # Determinar el número del semestre basado en las fechas
            if fecha_inicio_obj.month <= 6 and fecha_fin_obj.month <= 6:
                semestre = 1
            else:
                semestre = 2

            # Generar código automáticamente basado en el año y semestre
            codigo_generado = f"{año_actual}-{semestre}"

            # Validar código proporcionado si existe
            codigo_proporcionado = data.get("codigo", "")
            if codigo_proporcionado and codigo_proporcionado != codigo_generado:
                return JsonResponse({
                    "mensaje": f"El código debe ser {codigo_generado} basado en las fechas proporcionadas",
                    "codigo_sugerido": codigo_generado
                }, status=400)

            # Usar el código generado o el proporcionado si es correcto
            codigo_final = codigo_proporcionado if codigo_proporcionado else codigo_generado

            # Verificar si ya existe un periodo con el mismo código
            if Periodo.objects.filter(codigo=codigo_final).exists():
                return JsonResponse({
                    "mensaje": f"Ya existe un periodo con el código {codigo_final}, puedes editarlo"
                }, status=400)

            # Crear y guardar el nuevo periodo
            nuevo_periodo = Periodo(
                codigo=codigo_final,
                fecha_inicio=fecha_inicio_obj,
                fecha_fin=fecha_fin_obj,
                activo=data.get("activo", True)
            )
            nuevo_periodo.save()

            return JsonResponse({
                "mensaje": "Periodo creado exitosamente",
                "id": nuevo_periodo.id
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)

    return JsonResponse({"mensaje": "Método no permitido"}, status=405)


@csrf_exempt
def detalle_periodo(request, periodo_id):
    """Obtiene, actualiza o elimina un periodo específico"""
    if not request.user.is_authenticated or request.user.rol not in ["Administrador", "Director"]:
        return JsonResponse({"mensaje": "No autorizado"}, status=403)
    
    try:
        periodo = Periodo.objects.get(id=periodo_id)
    except ObjectDoesNotExist:
        return JsonResponse({"mensaje": "Periodo no encontrado"}, status=404)
    
    # GET: Obtener detalle de un periodo
    if request.method == "GET":
        data = {
            "id": periodo.id,
            "codigo": periodo.codigo,
            "fecha_inicio": periodo.fecha_inicio,
            "fecha_fin": periodo.fecha_fin,
            "activo": periodo.activo
        }
        return JsonResponse(data)
    
    # PUT: Actualizar un periodo
    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            fecha_actual = datetime.now().date()
            año_actual = fecha_actual.year
            
            # Verificar si hay un periodo activo y en curso (que no sea este mismo)
            periodo_activo = Periodo.objects.filter(
                activo=True, 
                fecha_fin__gte=fecha_actual
            ).exclude(id=periodo_id).first()
            
            # Si existe un periodo activo y en curso, solo permitir editar si:
            # 1. Este es el periodo activo actual, o
            # 2. Se está desactivando el periodo actual
            if periodo_activo and data.get("activo", False) is True and not periodo.activo:
                return JsonResponse({
                    "mensaje": "No se puede activar este periodo mientras exista un periodo activo en curso",
                    "periodo_vigente": {
                        "codigo": periodo_activo.codigo,
                        "fecha_inicio": periodo_activo.fecha_inicio,
                        "fecha_fin": periodo_activo.fecha_fin
                    }
                }, status=400)
        
             # Verificar si no hay cambios
            cambios_realizados = False
        
            # Modificar esta parte para permitir la edición sin cambios en fechas
            solo_cambio_estado = "activo" in data and (
                len(data) == 1 or 
                (len(data) > 1 and 
                (("fecha_inicio" not in data or data["fecha_inicio"] == periodo.fecha_inicio.strftime("%Y-%m-%d")) and
                ("fecha_fin" not in data or data["fecha_fin"] == periodo.fecha_fin.strftime("%Y-%m-%d"))))
            )
            
            # Validar formato del código si se está cambiando
            if "codigo" in data:
                cambios_realizados = True
                codigo = data["codigo"]
                if not re.match(r"^\d{4}-[12]$", codigo):
                    return JsonResponse({"mensaje": "El código debe tener el formato AAAA-S donde S es 1 o 2"}, status=400)
                
                # Verificar si el nuevo código ya existe y no es del periodo actual
                if Periodo.objects.filter(codigo=codigo).exclude(id=periodo_id).exists():
                    return JsonResponse({"mensaje": f"Ya existe un periodo con el código {codigo}"}, status=400)
                
                if periodo.codigo != codigo:
                    periodo.codigo = codigo
                    cambios_realizados = True
            
            # Actualizar otros campos
            if "fecha_inicio" in data:
                fecha_inicio_obj = datetime.strptime(data["fecha_inicio"], "%Y-%m-%d").date()
    
                 # Verificar si la fecha de inicio realmente cambió
                fecha_cambio = periodo.fecha_inicio != fecha_inicio_obj
    
                # Si solo estamos cambiando a inactivo, no validamos fechas
                if not solo_cambio_estado and fecha_cambio:
                    # Validar que la fecha de inicio sea del año actual y no sea pasada
                    if fecha_inicio_obj.year != año_actual or fecha_inicio_obj < fecha_actual:
                        return JsonResponse({
                            "mensaje": "La fecha de inicio debe ser del año actual y no puede ser anterior a la fecha actual"
                        }, status=400)
                
                if periodo.fecha_inicio != fecha_inicio_obj:
                    periodo.fecha_inicio = fecha_inicio_obj
                    cambios_realizados = True
            
            if "fecha_fin" in data:
                fecha_fin_obj = datetime.strptime(data["fecha_fin"], "%Y-%m-%d").date()
                
                # Verificar si la fecha de inicio realmente cambió
                fecha_cambio2 = periodo.fecha_fin != fecha_fin_obj
                
                # Si solo estamos cambiando a inactivo, no validamos fechas
                if not solo_cambio_estado and fecha_cambio2:
                    # Validar que la fecha de fin sea del año actual y no sea pasada
                    if fecha_fin_obj.year != año_actual or fecha_fin_obj < fecha_actual:
                        return JsonResponse({
                            "mensaje": "La fecha de fin debe ser del año actual y no puede ser anterior a la fecha actual"
                        }, status=400)
                    
                    # Validar que la fecha de inicio sea anterior a la fecha de fin
                    fecha_inicio_para_comparar = fecha_inicio_obj if "fecha_inicio" in data else periodo.fecha_inicio
                    if fecha_fin_obj < fecha_inicio_para_comparar:
                        return JsonResponse({
                            "mensaje": "La fecha de fin debe ser posterior a la fecha de inicio"
                        }, status=400)
                
                if periodo.fecha_fin != fecha_fin_obj:
                    periodo.fecha_fin = fecha_fin_obj
                    cambios_realizados = True
            
            if "activo" in data:
                # Si se está activando un periodo, verificar que no haya otro periodo activo
                if data["activo"] is True and not periodo.activo:
                    periodo_activo_existente = Periodo.objects.filter(
                        activo=True, 
                        fecha_fin__gte=fecha_actual
                    ).exclude(id=periodo_id).first()
                    
                    if periodo_activo_existente:
                        return JsonResponse({
                            "mensaje": "No se puede activar este periodo mientras exista otro periodo activo en curso",
                            "periodo_vigente": {
                                "codigo": periodo_activo_existente.codigo,
                                "fecha_inicio": periodo_activo_existente.fecha_inicio,
                                "fecha_fin": periodo_activo_existente.fecha_fin
                            }
                        }, status=400)
                
                if periodo.activo != data["activo"]:
                    periodo.activo = data["activo"]
                    cambios_realizados = True
            
            # Verificar si se realizó algún cambio
            if not cambios_realizados:
                return JsonResponse({"mensaje": "No se ha realizado ninguna modificación"}, status=200)
            
            periodo.save()
            return JsonResponse({"mensaje": "Periodo actualizado exitosamente"})
            
        except json.JSONDecodeError:
            return JsonResponse({"mensaje": "Error en el formato de los datos"}, status=400)
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=500)
    
    # DELETE: Eliminar un periodo
    elif request.method == "DELETE":
        try:
            # Verificar si el periodo tiene horarios asociados antes de eliminar
            if hasattr(periodo, 'horario_set') and periodo.horario_set.exists():
                return JsonResponse({
                    "mensaje": "No se puede eliminar este periodo porque tiene horarios asociados"
                }, status=400)
            
            # Verificar si el periodo está activo y en curso
            fecha_actual = datetime.now().date()
            if periodo.activo and periodo.fecha_fin >= fecha_actual:
                return JsonResponse({
                    "mensaje": "No se puede eliminar un periodo activo en curso. Debe desactivarlo primero."
                }, status=400)
            
            periodo.delete()
            return JsonResponse({"mensaje": "Periodo eliminado exitosamente"})
        except Exception as e:
            return JsonResponse({"mensaje": f"Error al eliminar: {str(e)}"}, status=500)
    
    return JsonResponse({"mensaje": "Método no permitido"}, status=405)

def periodo_vigente(request):
    hoy = timezone.now().date()
    try:
        periodo = Periodo.objects.filter(fecha_fin__gte=hoy, activo=True).order_by('fecha_inicio').first()
        if periodo:
            return JsonResponse({
                'codigo': periodo.codigo,
                'fecha_inicio': periodo.fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': periodo.fecha_fin.strftime('%Y-%m-%d'),
                'activo': periodo.activo,
                'id': periodo.id
            })
        return JsonResponse({'mensaje': 'No hay periodo vigente'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
def obtener_asignaturas_por_semestre(request):
    """Obtiene las asignaturas disponibles para el semestre del estudiante"""
    if request.method == 'GET':
        # Verificar que el usuario sea un estudiante
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Obtener el semestre del estudiante
        semestre = request.user.semestre
        if not semestre:
            return JsonResponse({'error': 'Estudiante sin semestre asignado'}, status=400)
        
        # Obtener asignaturas del semestre
        asignaturas = Asignatura.objects.filter(semestre=semestre)
        
        # Formatear respuesta
        asignaturas_data = []
        for asignatura in asignaturas:
            asignaturas_data.append({
                'id': asignatura.id,
                'nombre': asignatura.nombre,
                'area': asignatura.area.nombre,
                'semestre': asignatura.get_semestre_display()
            })
        
        return JsonResponse({'asignaturas': asignaturas_data})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def obtener_docentes_por_asignatura(request, asignatura_id):
    """Obtiene los docentes disponibles para una asignatura específica"""
    if request.method == 'GET':
        # Verificar que el usuario sea un estudiante
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        try:
            # Obtener la asignatura
            asignatura = Asignatura.objects.get(id=asignatura_id)
            
            # Obtener docentes según configuración
            docentes = asignatura.get_docentes()
            
            # Formatear respuesta
            docentes_data = []
            for docente in docentes:
                docentes_data.append({
                    'id': docente.id,
                    'correo': docente.correo,
                    'nombre_usuario': docente.nombre_usuario,
                    'subtipo': docente.subtipo_docente
                })
            
            return JsonResponse({'docentes': docentes_data})
        
        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Asignatura no encontrada'}, status=404)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def obtener_horarios_disponibles(request, docente_id):
    """Obtiene los bloques de horario disponibles para un docente específico"""
    if request.method == 'GET':
        # Verificar que el usuario sea un estudiante
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        try:
            # Obtener docente
            docente = Usuario.objects.get(id=docente_id, rol='Docente')
            
            # Obtener periodo activo
            try:
                periodo_activo = Periodo.objects.get(activo=True)
            except ObjectDoesNotExist:
                return JsonResponse({'error': 'No hay periodo académico activo'}, status=400)
            
            # Obtener la asignatura para la que se busca asesoría (opcional)
            asignatura_id = request.GET.get('asignatura_id')
            
            # Obtener el semestre del estudiante que hace la solicitud
            semestre_estudiante = request.user.semestre
            
            # Obtener fecha y hora actual
            ahora = timezone.localtime(timezone.now())
            fecha_actual = ahora.date()
            hora_actual = ahora.time()
            
            # Obtener horarios futuros (a partir de ahora)
            horarios = Horario.objects.filter(
                docente=docente,
                periodo=periodo_activo.codigo
            )
            
            # Obtener bloques de horario
            bloques_data = []
            for horario in horarios:
                bloques = BloqueHorario.objects.filter(horario=horario)
                
                # Filtrar bloques que sean futuros
                for bloque in bloques:
                    bloque_fecha = bloque.fecha
                    bloque_hora_fin = bloque.hora_fin
                    
                    # Solo incluir horarios futuros y que no estén en curso o por terminar
                    if bloque_fecha > fecha_actual or (bloque_fecha == fecha_actual and bloque_hora_fin > hora_actual):
                        # Verificar que el horario no haya comenzado o esté en curso
                        bloque_hora_inicio = bloque.hora_inicio
                        
                        # No mostrar bloques que ya han comenzado o están en curso
                        if bloque_fecha > fecha_actual or (bloque_fecha == fecha_actual and bloque_hora_inicio > hora_actual):
                            # Verificar si ya existen asesorías para este bloque específico
                            asesorias_existentes = Asesoria.objects.filter(
                                bloques_horario=bloque
                            ).exclude(
                                estado__in=['Cancelada', 'Rechazada']
                            )
                            
                            # Verificar si el estudiante actual ya tiene una asesoría en este bloque
                            tiene_asesoria = asesorias_existentes.filter(estudiante=request.user).exists()
                            
                            if asesorias_existentes.exists() and not tiene_asesoria:
                                # Obtener información de la asesoría existente
                                asesoria_existente = asesorias_existentes.first()
                                asignatura_existente_id = asesoria_existente.asignatura.id
                                semestre_estudiante_existente = asesoria_existente.estudiante.semestre
                                
                                # Mostrar bloque si:
                                # 1. Es para la misma asignatura, o
                                # 2. El estudiante existente es del mismo semestre
                                if (asignatura_id and int(asignatura_id) == asignatura_existente_id) or \
                                   (semestre_estudiante == semestre_estudiante_existente):
                                    bloques_data.append({
                                        'id': bloque.id,
                                        'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                                        'dia_semana': horario.dia,
                                        'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                                        'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                                        'lugar': bloque.lugar,  # Incluir el lugar en la respuesta
                                        'asignatura_asignada': asignatura_existente_id,
                                        'mismo_semestre': True,
                                        'disponible': True
                                    })
                            elif not tiene_asesoria:
                                # Si no hay asesorías existentes y el estudiante no tiene reserva en este bloque
                                bloques_data.append({
                                    'id': bloque.id,
                                    'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                                    'dia_semana': horario.dia,
                                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                                    'lugar': bloque.lugar,  # Incluir el lugar en la respuesta
                                    'asignatura_asignada': None,
                                    'mismo_semestre': False,
                                    'disponible': True
                                })
                            # Si tiene_asesoria es True, no se incluye el bloque en los resultados

            return JsonResponse({'bloques_horario': bloques_data})

        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Docente no encontrado'}, status=404)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def solicitar_asesoria(request):
    """Crea una nueva solicitud de asesoría con uno o más bloques de horario"""
    if request.method == 'POST':
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        try:
            data = json.loads(request.body)
            asignatura_id = data.get('asignatura_id')
            docente_id = data.get('docente_id')
            bloques_horario_ids = data.get('bloques_horario_ids', [])  # Ahora es una lista
            motivo = data.get('motivo', '')
            
            # Asegurar compatibilidad con el formato anterior
            if 'bloque_horario_id' in data and not bloques_horario_ids:
                bloques_horario_ids = [data.get('bloque_horario_id')]
            
            if not all([asignatura_id, docente_id]) or not bloques_horario_ids:
                return JsonResponse({'error': 'Faltan datos obligatorios'}, status=400)
            
            # Verificar que todos los bloques sean del mismo día
            bloques_horario = []
            fecha_referencia = None
            
            for bloque_id in bloques_horario_ids:
                try:
                    bloque = BloqueHorario.objects.get(id=bloque_id)
                    bloques_horario.append(bloque)
                    
                    if fecha_referencia is None:
                        fecha_referencia = bloque.fecha
                    elif fecha_referencia != bloque.fecha:
                        return JsonResponse({
                            'error': 'Todos los bloques de horario deben ser del mismo día'
                        }, status=400)
                except ObjectDoesNotExist:
                    return JsonResponse({
                        'error': f'Bloque de horario con ID {bloque_id} no encontrado'
                    }, status=404)
            
            # Bloque principal (para mantener compatibilidad)
            bloque_principal = bloques_horario[0]
            
            # Validar datos básicos
            try:
                asignatura = Asignatura.objects.get(id=asignatura_id)
                docente = Usuario.objects.get(id=docente_id, rol='Docente')
                
                # Verificar que todos los bloques correspondan al docente seleccionado
                for bloque in bloques_horario:
                    if bloque.horario.docente.id != docente.id:
                        return JsonResponse({
                            'error': f'El bloque de horario {bloque.id} no corresponde al docente seleccionado'
                        }, status=400)
            except ObjectDoesNotExist as e:
                return JsonResponse({'error': str(e)}, status=404)
            
            # Verificar que no haya conflictos con la asignatura
            for bloque in bloques_horario:
                asesorias_existentes = Asesoria.objects.filter(
                    bloques_horario=bloque
                ).exclude(
                    estado__in=['Cancelada', 'Rechazada']
                )
                
                if asesorias_existentes.exists():
                    # Verificar si ya existe una asesoría para otra asignatura
                    for asesoria in asesorias_existentes:
                        if asesoria.asignatura.id != int(asignatura_id):
                            return JsonResponse({
                                'error': f'El bloque {bloque.id} ya está asignado para otra asignatura',
                                'asignatura_actual': asesoria.asignatura.id
                            }, status=400)
                    
                    # Verificar si el estudiante actual ya tiene asesoría en este bloque
                    if asesorias_existentes.filter(estudiante=request.user).exists():
                        return JsonResponse({
                            'error': f'Ya has solicitado una asesoría para el bloque {bloque.id}'
                        }, status=400)
            
            # Todo está correcto, proceder a crear la asesoría
            with transaction.atomic():
                # Crear la asesoría con el bloque principal
                asesoria = Asesoria.objects.create(
                    estudiante=request.user,
                    docente=docente,
                    asignatura=asignatura,
                    bloque_horario=bloque_principal,  # Para compatibilidad
                    motivo=motivo,
                    estado='Aprobada'
                )
                
                # Agregar todos los bloques a la asesoría
                for bloque in bloques_horario:
                    # El primer bloque ya se añade automáticamente en el save() de Asesoria
                    if bloque != bloque_principal:
                        AsesoriaBloque.objects.create(
                            asesoria=asesoria,
                            bloque_horario=bloque
                        )
            
            # Enviar notificación por correo
            try:
                # Formar la lista de bloques para el correo incluyendo el lugar
                bloques_texto = "\n".join([
                    f"- {bloque.fecha.strftime('%d/%m/%Y')} de {bloque.hora_inicio.strftime('%H:%M')} a {bloque.hora_fin.strftime('%H:%M')} | Lugar: {bloque.lugar}"
                    for bloque in bloques_horario
                ])
                
                send_mail(
                    'Nueva asesoría programada',
                    f'El estudiante {request.user.correo} ha programado una asesoría para la asignatura {asignatura.nombre}.\n'
                    f'Bloques horarios:\n{bloques_texto}\n'
                    f'Motivo: {motivo}',
                    'sistema@example.com',
                    [docente.correo],
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Error al enviar correo: {e}")
            
            return JsonResponse({
                'success': True,
                'message': 'Asesoría programada exitosamente',
                'asesoria_id': asesoria.id
            })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def cancelar_asesoria(request, asesoria_id):
    """Elimina una asesoría previamente solicitada"""
    if request.method == 'POST':
        # Verificar que el usuario sea un estudiante
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        try:
            # Obtener la asesoría
            asesoria = Asesoria.objects.get(id=asesoria_id, estudiante=request.user)
            
            # Verificar que la asesoría esté en un estado cancelable (no finalizada)
            if asesoria.estado not in ['Solicitada', 'Aprobada']:
                return JsonResponse({'error': 'No se puede cancelar una asesoría en estado ' + asesoria.estado}, status=400)
            
            # Obtener motivo de cancelación para notificación
            data = json.loads(request.body)
            
            # Guardar información del docente para notificación antes de eliminar
            docente_correo = asesoria.docente.correo
            estudiante_correo = asesoria.estudiante.correo
            asignatura = asesoria.asignatura.nombre
            
            # Obtener información de los bloques antes de eliminar
            bloques_info = [
                {
                    'fecha': bloque.fecha.strftime('%d/%m/%Y'),
                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                    'lugar': bloque.lugar  # Agregamos el lugar
                }
                for bloque in asesoria.bloques_horario.all()
            ]
            
            # Eliminar la asesoría (esto también eliminará las relaciones AsesoriaBloque)
            asesoria.delete()
            
            # Notificar al docente
            try:
                # Formar la lista de bloques para el correo incluyendo el lugar
                bloques_texto = "\n".join([
                    f"- {bloque['fecha']} de {bloque['hora_inicio']} a {bloque['hora_fin']} | Lugar: {bloque['lugar']}"
                    for bloque in bloques_info
                ])
                
                send_mail(
                    'Asesoría cancelada',
                    f'El estudiante {estudiante_correo} ha cancelado la asesoría de {asignatura} programada para:\n'
                    f'{bloques_texto}',
                    'sistema@example.com',
                    [docente_correo],
                    fail_silently=True,
                )
            except Exception as e:
                # Registrar el error pero continuar
                print(f"Error al enviar correo: {e}")
            
            return JsonResponse({
                'success': True,
                'message': 'Asesoría cancelada exitosamente'
            })
            
        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Asesoría no encontrada'}, status=404)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def mis_asesorias(request):
    if request.method == 'GET':
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Primero actualizamos todas las asesorías pasadas
        ahora = timezone.localtime(timezone.now())
        fecha_actual = ahora.date()
        hora_actual = ahora.time()
        
        # Obtener asesorías que necesitan actualización
        asesorias_para_actualizar = Asesoria.objects.filter(
            estudiante=request.user,
            estado__in=['Aprobada', 'En curso'],
            bloques_horario__fecha__lte=fecha_actual
        ).distinct()
        
        for asesoria in asesorias_para_actualizar:
            # Obtener el último bloque de la asesoría
            ultimo_bloque = asesoria.bloques_horario.order_by('-fecha', '-hora_fin').first()
            
            if ultimo_bloque:
                # Verificar si la asesoría ya finalizó
                if (ultimo_bloque.fecha < fecha_actual) or \
                   (ultimo_bloque.fecha == fecha_actual and ultimo_bloque.hora_fin < hora_actual):
                    asesoria.estado = 'Finalizada'
                    asesoria.save()
                # Verificar si está en curso
                elif (ultimo_bloque.fecha == fecha_actual and 
                      ultimo_bloque.hora_inicio <= hora_actual <= ultimo_bloque.hora_fin):
                    asesoria.estado = 'En curso'
                    asesoria.save()
        
        # Verificar si el estudiante tiene asesorías pendientes de calificar
        pendientes = Asesoria.objects.filter(
            estudiante=request.user,
            estado='Finalizada',
            calificacion__isnull=True
        ).count()
        
        # Incluir esta información en la respuesta
        response_data = {
            'asesorias': [],  # se llenará más abajo
            'calificaciones_pendientes': pendientes
        }
        
        asesorias = Asesoria.objects.filter(estudiante=request.user).exclude(estado='Cancelada').order_by('-fecha_solicitud')
        
        asesorias_data = []
        for asesoria in asesorias:
            # Obtenemos todos los bloques de la asesoría
            bloques = asesoria.bloques_horario.all().order_by('fecha', 'hora_inicio')
            
            # Obtenemos el primer y último bloque para determinar rango completo
            if not bloques.exists():
                continue  # Saltar asesorías sin bloques (no debería ocurrir)
                
            primer_bloque = bloques.first()
            ultimo_bloque = bloques.last() if bloques.count() > 1 else primer_bloque
            
            fecha_asesoria = primer_bloque.fecha
            hora_inicio = primer_bloque.hora_inicio
            hora_fin = ultimo_bloque.hora_fin
            
            # Determinar estados temporales
            ha_finalizado = (fecha_asesoria < fecha_actual) or \
                          (fecha_asesoria == fecha_actual and hora_fin < hora_actual)
            
            en_curso = (fecha_asesoria == fecha_actual) and \
                      (hora_inicio <= hora_actual <= hora_fin)
            
            # Lógica de actualización de estados en la base de datos
            if en_curso and asesoria.estado != "En curso":
                asesoria.estado = "En curso"
                asesoria.save()
            elif ha_finalizado and asesoria.estado != "Finalizada":
                asesoria.estado = "Finalizada"
                asesoria.save()
            elif not ha_finalizado and not en_curso and asesoria.estado not in ["Aprobada", "Solicitada"]:
                asesoria.estado = "Aprobada"
                asesoria.save()
            
            # Determinar qué estado mostrar (priorizando estados temporales)
            estado_mostrar = asesoria.estado
            if en_curso:
                estado_mostrar = "En curso"
            
            # Formatear la información de todos los bloques
            bloques_data = []
            for bloque in bloques:
                # Obtenemos la información de asistencia del docente para este bloque
                asesoria_bloque = AsesoriaBloque.objects.filter(asesoria=asesoria, bloque_horario=bloque).first()
                docente_asistio = asesoria_bloque.docente_asistio if asesoria_bloque else False
                
                bloques_data.append({
                    'id': bloque.id,
                    'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                    'lugar': bloque.lugar,  # Añadimos el lugar de cada bloque
                    'docente_asistio': docente_asistio  # Agregamos información de asistencia del docente
                })
            
            # Obtener el lugar del primer bloque para mostrar en la información principal
            lugar_principal = primer_bloque.lugar if primer_bloque else 'No especificado'
            
            # Verificar si todos los bloques tienen registro de asistencia del docente
            docente_asistio_todos = all(bloque.get('docente_asistio', False) for bloque in bloques_data)
            
            asesorias_data.append({
                'id': asesoria.id,
                'asignatura': asesoria.asignatura.nombre,
                'docente': asesoria.docente.correo,
                'nombre_docente': asesoria.docente.nombre_usuario,
                'fecha': fecha_asesoria.strftime('%Y-%m-%d'),
                'hora_inicio': hora_inicio.strftime('%H:%M'),
                'hora_fin': hora_fin.strftime('%H:%M'),
                'lugar': lugar_principal,  # Añadimos el lugar del bloque principal
                'bloques': bloques_data,
                'estado': estado_mostrar,
                'estado_original': asesoria.estado,
                'fecha_solicitud': asesoria.fecha_solicitud.strftime('%Y-%m-%d %H:%M'),
                'motivo': asesoria.motivo,
                'comentarios': asesoria.comentarios,
                'puede_calificar': estado_mostrar == 'Finalizada' and asesoria.calificacion is None and docente_asistio_todos,
                'puede_registrar_asistencia': estado_mostrar == 'Finalizada' and not all(bloque.get('asistencia_registrada', False) for bloque in bloques_data),
                'docente_asistio': docente_asistio_todos,  # Indicamos si el docente asistió a todos los bloques
                'calificacion': asesoria.calificacion,
                'comentario_calificacion': asesoria.comentario_calificacion,
                'ha_finalizado': ha_finalizado,
                'en_curso': en_curso,
                'tiene_calificacion_pendiente': asesoria.estado == 'Finalizada' and asesoria.calificacion is None and docente_asistio_todos,
            })
            
            response_data['asesorias'] = asesorias_data
        
        return JsonResponse(response_data)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def calificar_asesoria(request, asesoria_id):
    """Permite calificar una asesoría finalizada y notifica al docente"""
    if request.method == 'POST':
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        finalizar_asesorias_pasadas()
        
        try:
            asesoria = Asesoria.objects.get(id=asesoria_id, estudiante=request.user)
            
            if asesoria.estado != 'Finalizada':
                return JsonResponse({'error': 'Solo se pueden calificar asesorías finalizadas'}, status=400)
            
            if asesoria.calificacion is not None:
                return JsonResponse({'error': 'Esta asesoría ya fue calificada'}, status=400)
            
            # Verificar si todos los bloques tienen registro de asistencia del docente como "asistió"
            bloques = asesoria.bloques_asesoria.all()
            docente_asistio_todos = all(bloque.docente_asistio for bloque in bloques)
            
            if not docente_asistio_todos:
                return JsonResponse({'error': 'No se puede calificar la asesoría porque el docente no asistió a todos los bloques'}, status=400)
            
            data = json.loads(request.body)
            calificacion = data.get('calificacion')
            comentario = data.get('comentario', '')
            
            if not calificacion or int(calificacion) not in range(1, 6):
                return JsonResponse({'error': 'La calificación debe ser un número entre 1 y 5'}, status=400)
            
            # Actualizar la asesoría
            asesoria.calificacion = calificacion
            asesoria.comentario_calificacion = comentario
            asesoria.fecha_calificacion = timezone.now()
            asesoria.save()
            
            # Obtener el lugar para el correo
            bloque_principal = asesoria.bloque_horario
            lugar = bloque_principal.lugar if bloque_principal else 'No especificado'
            
            # Enviar notificación por correo al docente
            email_sent = False
            try:
                subject = 'Tu asesoría ha sido calificada'
                message = f"""
                El estudiante {request.user.get_full_name() or request.user.correo} ha calificado tu asesoría:
                
                Asignatura: {asesoria.asignatura.nombre}
                Fecha: {asesoria.bloque_horario.fecha.strftime('%d/%m/%Y')}
                Hora: {asesoria.bloque_horario.hora_inicio.strftime('%H:%M')} - {asesoria.bloque_horario.hora_fin.strftime('%H:%M')}
                Lugar: {lugar}
                Calificación: {'⭐' * int(calificacion)} ({calificacion}/5)
                Comentario: {comentario or 'Sin comentario'}
                """
                
                # Versión simple sin HTML para mayor compatibilidad
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [asesoria.docente.correo],
                    fail_silently=False,
                )
                email_sent = True
                
            except Exception as e:
                print(f"Error al enviar correo: {e}")
                # Puedes registrar este error en tus logs
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error enviando correo de calificación: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': 'Asesoría calificada exitosamente' + (' y docente notificado' if email_sent else ' (pero no se pudo notificar al docente)'),
            })
            
        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Asesoría no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def registrar_asistencia_docente(request, asesoria_id):
    """Permite al estudiante registrar si el docente asistió o no a los bloques de asesoría"""
    if request.method == 'POST':
        if not request.user.is_authenticated or request.user.rol != 'Estudiante':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        try:
            asesoria = Asesoria.objects.get(id=asesoria_id, estudiante=request.user)
            
            if asesoria.estado != 'Finalizada':
                return JsonResponse({'error': 'Solo se puede registrar asistencia en asesorías finalizadas'}, status=400)
            
            data = json.loads(request.body)
            bloques_asistencia = data.get('bloques_asistencia', [])
            
            if not bloques_asistencia:
                return JsonResponse({'error': 'No se proporcionó información de asistencia'}, status=400)
            
            # Actualizar la asistencia para cada bloque
            for bloque_data in bloques_asistencia:
                bloque_id = bloque_data.get('bloque_id')
                docente_asistio = bloque_data.get('docente_asistio', False)
                
                try:
                    # Buscar el bloque de asesoría correspondiente
                    asesoria_bloque = AsesoriaBloque.objects.get(
                        asesoria=asesoria,
                        bloque_horario_id=bloque_id
                    )
                    
                    # Actualizar el campo de asistencia del docente
                    asesoria_bloque.docente_asistio = docente_asistio
                    asesoria_bloque.save()
                    
                except AsesoriaBloque.DoesNotExist:
                    return JsonResponse({
                        'error': f'El bloque con ID {bloque_id} no está asociado a esta asesoría'
                    }, status=400)
            
            # Verificar si todos los bloques tienen registro de asistencia del docente
            bloques = asesoria.bloques_asesoria.all()
            docente_asistio_todos = all(bloque.docente_asistio for bloque in bloques)
            
            # Si el docente no asistió a algún bloque y hay calificación, eliminarla
            if not docente_asistio_todos and asesoria.calificacion is not None:
                asesoria.calificacion = None
                asesoria.comentario_calificacion = None
                asesoria.fecha_calificacion = None
                asesoria.save()
            
            # Mensaje apropiado basado en la asistencia
            if docente_asistio_todos:
                mensaje = "Asistencia registrada correctamente. Ahora puedes calificar la asesoría."
            else:
                mensaje = "Asistencia registrada. No se puede calificar la asesoría porque el docente no asistió a todos los bloques."
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'docente_asistio_todos': docente_asistio_todos,
                'puede_calificar': docente_asistio_todos
            })
            
        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Asesoría no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def asesorias_programadas_docente(request):
    """Obtiene las asesorías programadas para el docente actual"""
    if request.method == 'GET':
        # Verificar que el usuario sea un docente
        if not request.user.is_authenticated or request.user.rol != 'Docente':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # AÑADIDO: Primero actualizamos todas las asesorías pasadas
        finalizar_asesorias_pasadas()
        
        # Obtener periodo activo
        try:
            periodo_activo = Periodo.objects.get(activo=True)
        except ObjectDoesNotExist:
            return JsonResponse({'error': 'No hay periodo académico activo'}, status=400)
        
        # Obtener fecha y hora actual
        ahora = timezone.now()
        fecha_actual = ahora.date()
        hora_actual = ahora.time()
        
        # Filtrar por fecha (opcional)
        fecha = request.GET.get('fecha')
        
        # SOLUCIÓN: Modificamos la lógica de búsqueda para incluir correctamente las asesorías del día actual
        if fecha:
            try:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                
                # Si la fecha solicitada es hoy, incluir todas las asesorías de hoy sin importar la hora
                if fecha_obj == fecha_actual:
                    asesorias = Asesoria.objects.filter(
                        docente=request.user,
                        estado='Aprobada',
                        bloque_horario__fecha=fecha_obj
                    ).distinct()
                else:
                    # Para cualquier otra fecha, mostramos todas las asesorías de ese día
                    asesorias = Asesoria.objects.filter(
                        docente=request.user,
                        estado='Aprobada',
                        bloque_horario__fecha=fecha_obj
                    ).distinct()
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)
        else:
            # Si no se especifica fecha, mostrar TODAS las asesorías pendientes
            # INCLUIMOS EXPLÍCITAMENTE las del día actual sin filtrar por hora
            asesorias = Asesoria.objects.filter(
                docente=request.user,
                estado='Aprobada'
            ).filter(
                Q(bloque_horario__fecha__gt=fecha_actual) |  # Fechas futuras
                Q(bloque_horario__fecha=fecha_actual)        # Todas las del día actual
            ).distinct()
        
        # Ordenar por fecha y hora
        asesorias = asesorias.order_by('bloque_horario__fecha', 'bloque_horario__hora_inicio')
        
        # Formatear la respuesta
        asesorias_data = []
        for asesoria in asesorias:
            # Esto asume que la asesoria tiene un bloque_horario (singular)
            if not hasattr(asesoria, 'bloque_horario') or not asesoria.bloque_horario:
                continue
                
            # Datos del estudiante, con manejo seguro de campos opcionales
            estudiante_data = {
                'id': asesoria.estudiante.id,
                'correo': asesoria.estudiante.correo,
                'nombre_usuario': asesoria.estudiante.nombre_usuario,
                'semestre': getattr(asesoria.estudiante, 'semestre', None)
            }
            
            # Agregar campo grupo si existe
            if hasattr(asesoria.estudiante, 'grupo'):
                estudiante_data['grupo'] = asesoria.estudiante.grupo
            
            # Crear datos de la asesoria
            asesoria_data = {
                'id': asesoria.id,
                'estudiante': estudiante_data,
                'asignatura': asesoria.asignatura.nombre,
                'fecha': asesoria.bloque_horario.fecha.strftime('%Y-%m-%d'),
                'hora_inicio': asesoria.bloque_horario.hora_inicio.strftime('%H:%M'),
                'hora_fin': asesoria.bloque_horario.hora_fin.strftime('%H:%M'),
                'estado': asesoria.estado,
                'motivo': asesoria.motivo
            }
            
            # Si es una estructura con múltiples bloques, agregarlos
            if hasattr(asesoria, 'bloques_horario'):
                bloques = asesoria.bloques_horario.all().order_by('fecha', 'hora_inicio')
                if bloques.exists():
                    bloques_data = []
                    for bloque in bloques:
                        bloques_data.append({
                            'id': bloque.id,
                            'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                            'dia_semana': getattr(bloque, 'dia_semana', ''),
                            'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                            'hora_fin': bloque.hora_fin.strftime('%H:%M')
                        })
                    
                    
                    
                    asesoria_data['bloques'] = bloques_data
                    
            
            asesorias_data.append(asesoria_data)
        
        return JsonResponse({'asesorias': asesorias_data})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def asesorias_finalizadas_docente(request):
    """Obtiene las asesorías finalizadas para el docente actual con asistencia por bloque"""
    if request.method == 'GET':
        # Verificar que el usuario sea un docente
        if not request.user.is_authenticated or request.user.rol != 'Docente':
            return JsonResponse({'error': 'No autorizado'}, status=403)

        # AÑADIDO: Finalizar asesorías pasadas antes de realizar la consulta
        finalizar_asesorias_pasadas()
        
        # Obtener fecha actual para filtrar
        fecha_actual = timezone.now().date()
        
        # Filtrar asesorías finalizadas (estado Finalizada) y que pertenezcan al docente
        # Usamos prefetch_related para optimizar la consulta de bloques
        asesorias = Asesoria.objects.filter(
            docente=request.user,
            estado='Finalizada'
        ).prefetch_related(
            'bloques_asesoria__bloque_horario'
        ).order_by('-bloque_horario__fecha', '-bloque_horario__hora_inicio')

        # Filtrar por fecha si se proporciona
        fecha_filtro = request.GET.get('fecha')
        if fecha_filtro:
            try:
                fecha_obj = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
                asesorias = asesorias.filter(bloques_horario__fecha=fecha_obj)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)

        # Formatear respuesta
        asesorias_data = []
        for asesoria in asesorias:
            # Obtener todos los bloques de la asesoría con su información de asistencia
            asesorias_bloques = AsesoriaBloque.objects.filter(
                asesoria=asesoria
            ).select_related('bloque_horario').order_by('bloque_horario__fecha', 'bloque_horario__hora_inicio')
            
            # Solo continuar si hay bloques disponibles
            if not asesorias_bloques.exists():
                continue
                
            # Obtener el primer y último bloque para mostrar el rango general
            primer_asesoria_bloque = asesorias_bloques.first()
            ultimo_asesoria_bloque = asesorias_bloques.last()
            
            primer_bloque = primer_asesoria_bloque.bloque_horario
            ultimo_bloque = ultimo_asesoria_bloque.bloque_horario
            
            # Preparar datos de los bloques con información de asistencia
            bloques_data = []
            for asesoria_bloque in asesorias_bloques:
                bloque = asesoria_bloque.bloque_horario
                bloques_data.append({
                    'id': bloque.id,
                    'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                    'lugar': bloque.lugar,
                    'asistio': asesoria_bloque.asistio,
                    'temas_tratados': asesoria_bloque.temas_tratados,
                    'observaciones': asesoria_bloque.observaciones
                })
            
            # Determinar si es un bloque único o múltiple
            es_bloque_unico = len(bloques_data) == 1
            
            asesorias_data.append({
                'id': asesoria.id,
                'estudiante': {
                    'id': asesoria.estudiante.id,
                    'correo': asesoria.estudiante.correo,
                    'nombre_usuario': asesoria.estudiante.nombre_usuario,
                    'semestre': asesoria.estudiante.semestre,
                    'grupo': asesoria.estudiante.grupo
                },
                'asignatura': asesoria.asignatura.nombre,
                # Información del primer bloque (para compatibilidad)
                'fecha': primer_bloque.fecha.strftime('%Y-%m-%d'),
                'hora_inicio': primer_bloque.hora_inicio.strftime('%H:%M'),
                'hora_fin': primer_bloque.hora_fin.strftime('%H:%M'),
                'lugar': primer_bloque.lugar,
                # Información general de la asesoría
                'asistio': asesoria.asistio,  # Asistencia general
                'temas_tratados': asesoria.temas_tratados,
                'observaciones': asesoria.observaciones,
                'compromisos': asesoria.compromisos,
                'calificacion': asesoria.calificacion,
                'comentario_calificacion': asesoria.comentario_calificacion,
                # Información detallada de todos los bloques
                'bloques': bloques_data,
                # Nuevo campo para indicar si es un bloque único
                'es_bloque_unico': es_bloque_unico
            })

        return JsonResponse({'asesorias': asesorias_data})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
def registrar_asistencia_asesoria(request, asesoria_id):
    """Permite al docente registrar la asistencia y detalles de una asesoría finalizada, por cada bloque"""
    if request.method == 'POST':
        # Verificar que el usuario sea un docente
        if not request.user.is_authenticated or request.user.rol != 'Docente':
            return JsonResponse({'error': 'No autorizado'}, status=403)

        try:
            # Obtener la asesoría con todos sus bloques relacionados
            asesoria = Asesoria.objects.get(id=asesoria_id, docente=request.user)
            
            # Verificar que la asesoría esté finalizada
            if asesoria.estado != 'Finalizada':
                return JsonResponse({'error': 'Solo se puede registrar información en asesorías finalizadas'}, status=400)

            # Obtener datos del request
            data = json.loads(request.body)
            
            # Obtener todos los bloques de esta asesoría
            asesorias_bloques = AsesoriaBloque.objects.filter(asesoria=asesoria).select_related('bloque_horario')
            es_bloque_unico = asesorias_bloques.count() == 1
            
            # Para asesorías con un solo bloque, aplicamos la asistencia general directamente
            if es_bloque_unico:
                # Registrar datos generales de la asesoría
                if 'asistio' in data:
                    asesoria.asistio = data.get('asistio')
                    # También actualizar el único bloque con el mismo valor de asistencia
                    if asesorias_bloques.exists():
                        asesoria_bloque = asesorias_bloques.first()
                        asesoria_bloque.asistio = data.get('asistio')
                        asesoria_bloque.save()
                if 'temas_tratados' in data:
                    asesoria.temas_tratados = data.get('temas_tratados', '')
                    # También actualizar el único bloque con el mismo valor de temas tratados
                    if asesorias_bloques.exists():
                        asesoria_bloque = asesorias_bloques.first()
                        asesoria_bloque.temas_tratados = data.get('temas_tratados', '')
                        asesoria_bloque.save()
                if 'observaciones' in data:
                    asesoria.observaciones = data.get('observaciones', '')
                    # También actualizar el único bloque con el mismo valor de observaciones
                    if asesorias_bloques.exists():
                        asesoria_bloque = asesorias_bloques.first()
                        asesoria_bloque.observaciones = data.get('observaciones', '')
                        asesoria_bloque.save()
                if 'compromisos' in data:
                    asesoria.compromisos = data.get('compromisos', '')
                asesoria.save()
            else:
                # Para asesorías con múltiples bloques, la asistencia general se basa en la asistencia de los bloques
                # Registrar datos generales de la asesoría (compatibilidad)
                if 'temas_tratados' in data:
                    asesoria.temas_tratados = data.get('temas_tratados', '')
                if 'observaciones' in data:
                    asesoria.observaciones = data.get('observaciones', '')
                if 'compromisos' in data:
                    asesoria.compromisos = data.get('compromisos', '')
            
            # Procesar la asistencia por bloques (siempre procesamos los bloques)
            bloques_data = data.get('bloques', [])
            for bloque_data in bloques_data:
                bloque_id = bloque_data.get('id')
                asistio_bloque = bloque_data.get('asistio', False)
                temas_bloque = bloque_data.get('temas_tratados', '')
                observaciones_bloque = bloque_data.get('observaciones', '')
                
                try:
                    # Buscar si existe el bloque de asesoría
                    asesoria_bloque = AsesoriaBloque.objects.get(
                        asesoria=asesoria,
                        bloque_horario__id=bloque_id
                    )
                    
                    # Actualizar los campos de asistencia
                    asesoria_bloque.asistio = asistio_bloque
                    asesoria_bloque.temas_tratados = temas_bloque
                    asesoria_bloque.observaciones = observaciones_bloque
                    asesoria_bloque.save()
                except AsesoriaBloque.DoesNotExist:
                    # Si no existe, ignorar este bloque
                    pass
            
            # Para múltiples bloques, actualizamos la asistencia general basada en bloques
            if not es_bloque_unico:
                # Verificar si al menos un bloque tiene asistencia para actualizar el estado general
                asistio_general = AsesoriaBloque.objects.filter(asesoria=asesoria, asistio=True).exists()
                asesoria.asistio = asistio_general
                asesoria.save()
            
            # Obtener todos los bloques de horario actualizados para la respuesta
            asesorias_bloques = AsesoriaBloque.objects.filter(asesoria=asesoria).select_related('bloque_horario')
            
            # Preparar datos de los bloques para la respuesta
            bloques_data = []
            for asesoria_bloque in asesorias_bloques:
                bloque = asesoria_bloque.bloque_horario
                bloques_data.append({
                    'id': bloque.id,
                    'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                    'lugar': bloque.lugar,
                    'asistio': asesoria_bloque.asistio,
                    'temas_tratados': asesoria_bloque.temas_tratados,
                    'observaciones': asesoria_bloque.observaciones
                })

            return JsonResponse({
                'success': True,
                'message': 'Información de la asesoría actualizada exitosamente',
                'asesoria_id': asesoria.id,
                'bloques': bloques_data,
                'es_bloque_unico': es_bloque_unico
            })

        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Asesoría no encontrada o no pertenece a este docente'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'GET':
        # Para solicitudes GET, devolver información de la asesoría y sus bloques
        try:
            asesoria = Asesoria.objects.get(id=asesoria_id, docente=request.user)
            
            if asesoria.estado != 'Finalizada':
                return JsonResponse({'error': 'Solo se puede registrar información en asesorías finalizadas'}, status=400)

            # Obtener todos los bloques de horario asociados con su registro de asistencia
            asesorias_bloques = AsesoriaBloque.objects.filter(asesoria=asesoria).select_related('bloque_horario')
            es_bloque_unico = asesorias_bloques.count() == 1
            
            bloques_data = []
            for asesoria_bloque in asesorias_bloques:
                bloque = asesoria_bloque.bloque_horario
                bloques_data.append({
                    'id': bloque.id,
                    'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                    'lugar': bloque.lugar,
                    'asistio': asesoria_bloque.asistio,
                    'temas_tratados': asesoria_bloque.temas_tratados,
                    'observaciones': asesoria_bloque.observaciones
                })

            return JsonResponse({
                'asesoria_id': asesoria.id,
                'estado': asesoria.estado,
                'asistio': asesoria.asistio,  # Campo general para compatibilidad
                'temas_tratados': asesoria.temas_tratados,
                'observaciones': asesoria.observaciones,
                'compromisos': asesoria.compromisos,
                'bloques': bloques_data,
                'es_bloque_unico': es_bloque_unico
            })

        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Asesoría no encontrada o no pertenece a este docente'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)
def historial_asesorias(request):
    """
    Muestra el historial completo de asesorías con capacidad de filtrado
    y asistencia por bloque
    """
    if request.method == 'GET':
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Preparar la consulta base según el rol del usuario
        if request.user.rol in ['Director']:
            asesorias = Asesoria.objects.all()
            
            if request.user.rol == 'Director' and request.user.subtipo_director:
                areas = Area.objects.filter(subtipo_director=request.user.subtipo_director)
                asignaturas = Asignatura.objects.filter(area__in=areas)
                asesorias = asesorias.filter(asignatura__in=asignaturas)
                
        elif request.user.rol == 'Docente':
            asesorias = Asesoria.objects.filter(docente=request.user)
        elif request.user.rol == 'Estudiante':
            asesorias = Asesoria.objects.filter(estudiante=request.user)
        else:
            return JsonResponse({'error': 'Rol de usuario no válido'}, status=403)
        
        # Aplicar filtros si se proporcionan
        # Filtrar por fecha
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                asesorias = asesorias.filter(bloque_horario__fecha__gte=fecha_inicio_obj)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_inicio inválido. Use YYYY-MM-DD'}, status=400)
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                asesorias = asesorias.filter(bloque_horario__fecha__lte=fecha_fin_obj)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_fin inválido. Use YYYY-MM-DD'}, status=400)
        
        # Filtrar por asistencia (ahora verificamos asistencia por bloque también)
        asistio_param = request.GET.get('asistio')
        bloque_id = request.GET.get('bloque_id')
        if asistio_param is not None:
            try:
                asistio_value = asistio_param.lower() == 'true'
                
                if bloque_id:
                    try:
                        if asistio_value:
                            asesorias = asesorias.filter(bloques_asesoria__bloque_horario__id=bloque_id, 
                                                    bloques_asesoria__asistio=True)
                        else:
                         asesorias = asesorias.filter(bloques_asesoria__bloque_horario__id=bloque_id, 
                                                    bloques_asesoria__asistio=False)
                    except ValueError:
                        return JsonResponse({'error': 'ID de bloque inválido'}, status=400)
                else:
                # Filtro general por cualquier bloque con asistencia
                    if asistio_value:
                        asesorias = asesorias.filter(Q(bloques_asesoria__asistio=True)).distinct()
                    else:
                    # No asistió a ningún bloque
                        asesorias = asesorias.exclude(Q(bloques_asesoria__asistio=True))
            except ValueError:
                return JsonResponse({'error': 'Valor inválido para parámetro asistio. Use true o false'}, status=400)
        
        # Resto de filtros (código existente)...
        # (Manteniendo el código existente de los demás filtros)
        
        # Filtrar por asignatura
        asignatura_id = request.GET.get('asignatura_id')
        if asignatura_id:
            try:
                asesorias = asesorias.filter(asignatura_id=int(asignatura_id))
            except ValueError:
                return JsonResponse({'error': 'ID de asignatura inválido'}, status=400)
        
        # Filtrar por docente (solo para admin/director)
        docente_id = request.GET.get('docente_id')
        if docente_id and request.user.rol in ['Director']:
            try:
                asesorias = asesorias.filter(docente_id=int(docente_id))
            except ValueError:
                return JsonResponse({'error': 'ID de docente inválido'}, status=400)
        
        # Filtrar por estudiante (solo para admin/director/docente)
        estudiante_id = request.GET.get('estudiante_id')
        if estudiante_id and request.user.rol in ['Director', 'Docente']:
            try:
                asesorias = asesorias.filter(estudiante_id=int(estudiante_id))
            except ValueError:
                return JsonResponse({'error': 'ID de estudiante inválido'}, status=400)
        
        # Ordenar y paginar (código existente)...
        # (Manteniendo el código existente para ordenación y paginación)
        
        # Ordenar los resultados (por defecto, de más reciente a más antiguo)
        orden = request.GET.get('orden', '-fecha_solicitud')
        ordenes_validos = ['fecha_solicitud', '-fecha_solicitud', 'bloque_horario__fecha', 
                        '-bloque_horario__fecha', 'estado', '-estado', 'asignatura__nombre', 
                        '-asignatura__nombre']
        
        if orden in ordenes_validos:
            asesorias = asesorias.order_by(orden)
        else:
            asesorias = asesorias.order_by('-fecha_solicitud')
        
        # Paginación
        pagina = request.GET.get('pagina', 1)
        items_por_pagina = request.GET.get('items_por_pagina', 20)
        
        try:
            pagina = int(pagina)
            items_por_pagina = min(int(items_por_pagina), 100)  # Limitar a máximo 100 por página
        except ValueError:
            pagina = 1
            items_por_pagina = 20
        
        # Calcular índices para paginación manual
        indice_inicio = (pagina - 1) * items_por_pagina
        indice_fin = indice_inicio + items_por_pagina
        
        # Obtener el total antes de aplicar la paginación
        total_asesorias = asesorias.count()
        
        # Aplicar paginación
        asesorias = asesorias[indice_inicio:indice_fin]
        
        # Formatear respuesta con datos completos
        asesorias_data = []
        for asesoria in asesorias:
            # Obtener todos los bloques con su información de asistencia
            asesorias_bloques = AsesoriaBloque.objects.filter(
                asesoria=asesoria
            ).select_related('bloque_horario').order_by('bloque_horario__fecha', 'bloque_horario__hora_inicio')
            
            bloques_data = []
            
            for asesoria_bloque in asesorias_bloques:
                bloque = asesoria_bloque.bloque_horario
                bloques_data.append({
                    'id': bloque.id,
                    'fecha': bloque.fecha.strftime('%Y-%m-%d'),
                    'dia_semana': bloque.horario.dia if hasattr(bloque, 'horario') and bloque.horario else '',
                    'hora_inicio': bloque.hora_inicio.strftime('%H:%M'),
                    'hora_fin': bloque.hora_fin.strftime('%H:%M'),
                    'lugar': bloque.lugar,
                    # Datos de asistencia específicos del bloque
                    'asistio': asesoria_bloque.asistio,
                    'temas_tratados': asesoria_bloque.temas_tratados,
                    'observaciones': asesoria_bloque.observaciones
                })
            
            # Si no hay bloques, continuar con la siguiente asesoría
            if not bloques_data:
                continue
                
            # Determinar el rango de tiempo completo de la asesoría usando el primer y último bloque
            primer_bloque = asesorias_bloques.first().bloque_horario if asesorias_bloques.exists() else None
            ultimo_bloque = asesorias_bloques.last().bloque_horario if asesorias_bloques.count() > 1 else primer_bloque
            
            if not primer_bloque:
                continue
                
            # Obtener estado actual basado en fecha/hora
            ahora = timezone.localtime(timezone.now())
            fecha_actual = ahora.date()
            hora_actual = ahora.time()
            
            # Determine estado temporal (en curso, finalizada)
            ha_finalizado = (primer_bloque.fecha < fecha_actual) or \
                          (primer_bloque.fecha == fecha_actual and ultimo_bloque.hora_fin < hora_actual)
            
            en_curso = (primer_bloque.fecha == fecha_actual) and \
                      (primer_bloque.hora_inicio <= hora_actual <= ultimo_bloque.hora_fin)
            
            # Determinar qué estado mostrar (priorizando estados temporales)
            estado_mostrar = asesoria.estado
            if en_curso:
                estado_mostrar = "En curso"
            elif ha_finalizado and asesoria.estado != "Finalizada":
                estado_mostrar = "Finalizada"
            
            # El resto de la construcción de respuesta (manteniendo los campos existentes)
            asesoria_info = {
                'id': asesoria.id,
                'estudiante': {
                    'id': asesoria.estudiante.id,
                    'correo': asesoria.estudiante.correo,
                    'nombre': asesoria.estudiante.nombre_usuario,
                    'semestre': asesoria.estudiante.semestre,
                    'grupo': asesoria.estudiante.grupo
                },
                'docente': {
                    'id': asesoria.docente.id,
                    'correo': asesoria.docente.correo,
                    'nombre': asesoria.docente.nombre_usuario,
                    'subtipo_docente': asesoria.docente.subtipo_docente
                },
                'asignatura': {
                    'id': asesoria.asignatura.id,
                    'nombre': asesoria.asignatura.nombre,
                    'semestre': asesoria.asignatura.semestre,
                    'area': asesoria.asignatura.area.nombre
                },
                # Información del bloque principal (compatibilidad)
                'bloque_principal': {
                    'id': asesoria.bloque_horario.id,
                    'fecha': asesoria.bloque_horario.fecha.strftime('%Y-%m-%d'),
                    'hora_inicio': asesoria.bloque_horario.hora_inicio.strftime('%H:%M'),
                    'hora_fin': asesoria.bloque_horario.hora_fin.strftime('%H:%M'),
                    'lugar': asesoria.bloque_horario.lugar,
                    'asistio_principal': asesoria.asistio  # Para compatibilidad
                },
                # Información de todos los bloques con detalle de asistencia
                'bloques': bloques_data,
                # Información del rango completo
                'fecha_inicio': primer_bloque.fecha.strftime('%Y-%m-%d'),
                'hora_inicio_primer_bloque': primer_bloque.hora_inicio.strftime('%H:%M'),
                'hora_fin_ultimo_bloque': ultimo_bloque.hora_fin.strftime('%H:%M'),
                # Estados
                'estado': estado_mostrar,
                'estado_original': asesoria.estado,
                'ha_finalizado': ha_finalizado,
                'en_curso': en_curso,
                # Fechas
                'fecha_solicitud': asesoria.fecha_solicitud.strftime('%Y-%m-%d %H:%M'),
                # Detalles adicionales
                'motivo': asesoria.motivo,
                'comentarios': asesoria.comentarios,
                # Registro de asistencia general (compatibilidad)
                'asistio': asesoria.asistio,
                'temas_tratados': asesoria.temas_tratados,
                'observaciones': asesoria.observaciones,
                'compromisos': asesoria.compromisos,
                # Calificación
                'calificacion': asesoria.calificacion,
                'comentario_calificacion': asesoria.comentario_calificacion,
                'fecha_calificacion': asesoria.fecha_calificacion.strftime('%Y-%m-%d %H:%M') if asesoria.fecha_calificacion else None,
                'tiene_calificacion': asesoria.calificacion is not None,
                'es_un_solo_bloque': len(bloques_data) == 1,
            }
            
            asesorias_data.append(asesoria_info)
        
        # Construir respuesta final con paginación
        response = {
            'asesorias': asesorias_data,
            'paginacion': {
                'total': total_asesorias,
                'pagina_actual': pagina,
                'total_paginas': (total_asesorias + items_por_pagina - 1) // items_por_pagina,
                'items_por_pagina': items_por_pagina
            },
            # Incluir datos útiles para filtros en UI
            'filtros_disponibles': {
                'estados': [estado[0] for estado in Asesoria.ESTADOS],
            }
        }
        
        return JsonResponse(response)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def exportar_historial_asesorias(request):
    """
    Exporta el historial de asesorías a un archivo Excel con cada bloque como fila separada
    """
    if request.method == 'GET':
        if not request.user.is_authenticated or request.user.rol not in ['Director', 'Docente']:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Construir la consulta según el rol del usuario
        if request.user.rol == 'Director':
            asesorias = Asesoria.objects.all()
            
            if request.user.subtipo_director:
                areas = Area.objects.filter(subtipo_director=request.user.subtipo_director)
                asignaturas = Asignatura.objects.filter(area__in=areas)
                asesorias = asesorias.filter(asignatura__in=asignaturas)
                
        elif request.user.rol == 'Docente':
            asesorias = Asesoria.objects.filter(docente=request.user)
        else:
            return JsonResponse({'error': 'Rol no autorizado para exportar'}, status=403)
        
        # Aplicar filtros basados en parámetros GET
        # Filtrar por fecha
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                asesorias = asesorias.filter(bloque_horario__fecha__gte=fecha_inicio_obj)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_inicio inválido. Use YYYY-MM-DD'}, status=400)
        
        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                asesorias = asesorias.filter(bloque_horario__fecha__lte=fecha_fin_obj)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_fin inválido. Use YYYY-MM-DD'}, status=400)
        
        # Filtrar por estado de asistencia
        estado = request.GET.get('estado')
        if estado:
            if estado == 'asistio':
                asesorias = asesorias.filter(Q(bloques_asesoria__asistio=True)).distinct()
            elif estado == 'no_asistio':
                asesorias = asesorias.exclude(Q(bloques_asesoria__asistio=True))
        
        # Filtrar por asignatura
        asignatura_id = request.GET.get('asignatura_id')
        if asignatura_id:
            try:
                asesorias = asesorias.filter(asignatura_id=int(asignatura_id))
            except ValueError:
                return JsonResponse({'error': 'ID de asignatura inválido'}, status=400)
        
        # Filtrar por docente (solo para admin/director)
        docente_id = request.GET.get('docente_id')
        if docente_id and request.user.rol in ['Director']:
            try:
                asesorias = asesorias.filter(docente_id=int(docente_id))
            except ValueError:
                return JsonResponse({'error': 'ID de docente inválido'}, status=400)
        
        # Filtrar por estudiante (solo para admin/director/docente)
        estudiante_id = request.GET.get('estudiante_id')
        if estudiante_id and request.user.rol in ['Director', 'Docente']:
            try:
                asesorias = asesorias.filter(estudiante_id=int(estudiante_id))
            except ValueError:
                return JsonResponse({'error': 'ID de estudiante inválido'}, status=400)
        
        # Ordenar resultados
        asesorias = asesorias.order_by('-fecha_solicitud')
        
        # Crear un libro de Excel
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Asesorías"
        
        # Estilo para los encabezados
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Definir los encabezados (ajustados para bloques separados)
        encabezados = [
            'ID Asesoría', 'Fecha Solicitud', 'Estado', 
            'Estudiante', 'Correo Estudiante', 'Semestre', 'Grupo',
            'Docente', 'Correo Docente',
            'Asignatura', 'Área', 
            'Fecha Bloque', 'Hora Inicio', 'Hora Fin', 'Lugar',
            'Motivo', 'Asistió a este bloque', 'Temas Tratados', 'Observaciones', 
            'Compromisos', 'Calificación', 'Comentario Calificación',
            'Número de Bloque', 'Total Bloques'
        ]
        
        # Agregar encabezados
        for col_num, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = thin_border
        
        # Contador de filas
        row_num = 2
        
        # Agregar los datos
        for asesoria in asesorias:
            # Obtener todos los bloques de la asesoría ordenados
            asesorias_bloques = AsesoriaBloque.objects.filter(
                asesoria=asesoria
            ).select_related('bloque_horario').order_by('bloque_horario__fecha', 'bloque_horario__hora_inicio')
            
            cantidad_bloques = asesorias_bloques.count()
            
            # Datos comunes para todos los bloques de esta asesoría
            datos_comunes = [
                asesoria.id,
                asesoria.fecha_solicitud.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M'),
                asesoria.estado,
                asesoria.estudiante.nombre_usuario,
                asesoria.estudiante.correo,
                asesoria.estudiante.semestre,
                asesoria.estudiante.grupo,
                asesoria.docente.nombre_usuario,
                asesoria.docente.correo,
                asesoria.asignatura.nombre,
                asesoria.asignatura.area.nombre,
                # Los campos específicos del bloque se llenarán en el bucle
            ]
            
            # Para cada bloque, crear una fila en el Excel
            for i, asesoria_bloque in enumerate(asesorias_bloques, 1):
                bloque = asesoria_bloque.bloque_horario
                
                # Datos específicos de este bloque
                datos_bloque = [
                    bloque.fecha.strftime('%Y-%m-%d'),
                    bloque.hora_inicio.strftime('%H:%M'),
                    bloque.hora_fin.strftime('%H:%M'),
                    bloque.lugar,
                    asesoria.motivo,
                    'Sí' if asesoria_bloque.asistio else 'No',
                    asesoria_bloque.temas_tratados or '',
                    asesoria_bloque.observaciones or '',
                    asesoria.compromisos or '',
                    asesoria.calificacion if asesoria.calificacion is not None else '',
                    asesoria.comentario_calificacion or '',
                    i,  # Número de este bloque
                    cantidad_bloques  # Total de bloques
                ]
                
                # Combinar datos comunes y específicos del bloque
                row_data = datos_comunes + datos_bloque
                
                # Agregar la fila al Excel
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                row_num += 1
        
        # Ajustar el ancho de las columnas
        column_widths = {
            1: 10,   # ID Asesoría
            2: 18,   # Fecha Solicitud
            3: 12,   # Estado
            4: 25,   # Estudiante
            5: 30,   # Correo Estudiante
            6: 10,   # Semestre
            7: 10,   # Grupo
            8: 25,   # Docente
            9: 30,   # Correo Docente
            10: 25,  # Asignatura
            11: 20,  # Área
            12: 12,  # Fecha Bloque
            13: 10,  # Hora Inicio
            14: 10,  # Hora Fin
            15: 20,  # Lugar
            16: 30,  # Motivo
            17: 15,  # Asistió
            18: 40,  # Temas Tratados
            19: 40,  # Observaciones
            20: 40,  # Compromisos
            21: 12,  # Calificación
            22: 40,  # Comentario Calificación
            23: 12,  # Número de Bloque
            24: 12   # Total Bloques
        }
        
        for col_num, width in column_widths.items():
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = width
        
        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Historial_Asesorias_Detallado.xlsx"'
        
        # Guardar el libro de Excel en la respuesta
        wb.save(response)
        
        return response
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
def obtener_asesoria(request, asesoria_id):
    """
    Obtiene los detalles completos de una asesoría específica.
    Esta función es llamada desde MenuDocentes.jsx para mostrar el detalle de una asesoría.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    # Verificar que el usuario está autenticado
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
    
    try:
        # Obtener la asesoría
        asesoria = Asesoria.objects.get(id=asesoria_id)
        
        # Verificar permisos: debe ser el docente de la asesoría, el estudiante, o un admin/director
        es_docente = request.user == asesoria.docente
        es_estudiante = request.user == asesoria.estudiante
        es_admin = request.user.rol in ['Administrador', 'Director']
        
        if not (es_docente or es_estudiante or es_admin):
            return JsonResponse({'error': 'No tiene permisos para ver esta asesoría'}, status=403)
        
        # Obtener todos los bloques asociados a esta asesoría
        bloques_asesoria = AsesoriaBloque.objects.filter(asesoria=asesoria)
        bloques_data = []
        
        for bloque_asesoria in bloques_asesoria:
            bloque = bloque_asesoria.bloque_horario
            dia_semana = ''
            if bloque.fecha:
                dia_semana = bloque.fecha.strftime('%A')
                # Convertir nombres de días en inglés a español
                dias_semana = {
                    'Monday': 'Lunes',
                    'Tuesday': 'Martes',
                    'Wednesday': 'Miércoles',
                    'Thursday': 'Jueves',
                    'Friday': 'Viernes',
                    'Saturday': 'Sábado',
                    'Sunday': 'Domingo'
                }
                dia_semana = dias_semana.get(dia_semana, dia_semana)
            
            bloques_data.append({
                'id': bloque.id,
                'fecha': bloque.fecha.strftime('%d/%m/%Y') if bloque.fecha else '',
                'dia_semana': dia_semana,
                'hora_inicio': bloque.hora_inicio.strftime('%H:%M') if bloque.hora_inicio else '',
                'hora_fin': bloque.hora_fin.strftime('%H:%M') if bloque.hora_fin else '',
                'lugar': bloque.lugar or 'No especificado',
                'asistio': bloque_asesoria.asistio,
                'temas_tratados': bloque_asesoria.temas_tratados,
                'observaciones': bloque_asesoria.observaciones
            })
            
        asistio_todos = all(bloque.get('asistio', False) for bloque in bloques_data)
        asistio_alguno = any(bloque.get('asistio', False) for bloque in bloques_data)
        
        # Verificar si la asesoría ha finalizado (para UI)
        ha_finalizado = False
        if asesoria.estado == 'Finalizada':
            ha_finalizado = True
        else:
            # Comprobar si el último bloque ya pasó
            ultimo_bloque = bloques_asesoria.order_by('-bloque_horario__fecha', '-bloque_horario__hora_fin').first()
            if ultimo_bloque and ultimo_bloque.bloque_horario.fecha:
                fecha_hora_fin = datetime.combine(
                    ultimo_bloque.bloque_horario.fecha,
                    ultimo_bloque.bloque_horario.hora_fin
                )
                if fecha_hora_fin < timezone.now():
                    ha_finalizado = True
        
        # Verificar si tiene calificación
        tiene_calificacion = asesoria.calificacion is not None
        
        # Datos de la asignatura
        asignatura_data = {
            'id': asesoria.asignatura.id,
            'nombre': asesoria.asignatura.nombre,
            'semestre': asesoria.asignatura.semestre
        }
        
        # Datos del estudiante
        estudiante_data = {
            'id': asesoria.estudiante.id,
            'nombre': asesoria.estudiante.nombre_usuario,
            'correo': asesoria.estudiante.correo,
            'semestre': asesoria.estudiante.semestre,
            'grupo': asesoria.estudiante.grupo
        }
        
        # Datos del docente
        docente_data = {
            'id': asesoria.docente.id,
            'nombre': asesoria.docente.nombre_usuario,
            'correo': asesoria.docente.correo,
            'subtipo': asesoria.docente.subtipo_docente
        }
        
        # Construir objeto de respuesta completo
        asesoria_data = {
            'id': asesoria.id,
            'estado': asesoria.estado,
            'fecha_solicitud': asesoria.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            'motivo': asesoria.motivo,
            'comentarios': asesoria.comentarios,
            'ha_finalizado': ha_finalizado,
            'asistio': asesoria.asistio,
            'asistio_todos': asistio_todos,
            'asistio_alguno': asistio_alguno,
            'bloques': bloques_data,
            'temas_tratados': asesoria.temas_tratados,
            'observaciones': asesoria.observaciones,
            'compromisos': asesoria.compromisos,
            'tiene_calificacion': tiene_calificacion,
            'calificacion': asesoria.calificacion,
            'comentario_calificacion': asesoria.comentario_calificacion,
            'fecha_calificacion': asesoria.fecha_calificacion.strftime('%d/%m/%Y %H:%M') if asesoria.fecha_calificacion else None,
            'asignatura': asignatura_data,
            'estudiante': estudiante_data,
            'docente': docente_data,
            'bloques': bloques_data
        }
        
        return JsonResponse(asesoria_data)
        
    except ObjectDoesNotExist:
        return JsonResponse({'error': 'Asesoría no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



def finalizar_asesorias_pasadas():
    """
    Marca las asesorías aprobadas como finalizadas cuando su fecha y hora de fin han pasado.
    Considera tanto la fecha como la hora para determinar si una asesoría debe finalizarse.
    """
    ahora = timezone.localtime(timezone.now())  # Aseguramos usar zona horaria local
    fecha_actual = ahora.date()
    hora_actual = ahora.time()
    
    # Asesorías que deben finalizarse (con último bloque ya pasado)
    asesorias_para_actualizar = Asesoria.objects.filter(
        estado__in=['Aprobada', 'En curso'],
        bloques_horario__fecha__lte=fecha_actual
    ).distinct()
    
    for asesoria in asesorias_para_actualizar:
        # Obtener el último bloque de la asesoría
        ultimo_bloque = asesoria.bloques_horario.order_by('-fecha', '-hora_fin').first()
        
        if ultimo_bloque:
            # Verificar si la asesoría ya finalizó
            if (ultimo_bloque.fecha < fecha_actual) or \
               (ultimo_bloque.fecha == fecha_actual and ultimo_bloque.hora_fin < hora_actual):
                asesoria.estado = 'Finalizada'
                asesoria.save()
            # Verificar si está en curso
            elif (ultimo_bloque.fecha == fecha_actual and 
                  ultimo_bloque.hora_inicio <= hora_actual <= ultimo_bloque.hora_fin):
                asesoria.estado = 'En curso'
                asesoria.save()